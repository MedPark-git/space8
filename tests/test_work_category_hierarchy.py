import io
import importlib
import os
import unittest
from datetime import date

import sqlalchemy as sa
from alembic.migration import MigrationContext
from alembic.operations import Operations
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash


os.environ.setdefault("ALLOW_IN_MEMORY_DB", "1")
os.environ.setdefault("SESSION_COOKIE_SECURE", "0")
os.environ.setdefault(
    "BOOTSTRAP_ADMIN_PASSWORD_HASH",
    generate_password_hash("StartupAdmin123", method="scrypt"),
)

from app import (  # noqa: E402
    Department,
    Employee,
    Role,
    Task,
    TaskClassification,
    WorkCategory,
    create_app,
    db,
)


class WorkCategoryHierarchyTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": "sqlite://",
                "WTF_CSRF_ENABLED": False,
                "SESSION_COOKIE_SECURE": False,
            }
        )
        self.context = self.app.app_context()
        self.context.push()
        db.create_all()

        self.security = Department(name="보안전산팀", active=True)
        self.finance = Department(name="재무운영팀", active=True)
        self.admin_role = Role(
            name="관리자",
            data_scope="own",
            permissions={},
            is_system=True,
        )
        db.session.add_all([self.security, self.finance, self.admin_role])
        db.session.flush()
        self.admin = Employee(
            name="시스템 관리자",
            department=self.finance,
            position="팀장",
            login_id="admin1",
            password_hash="",
            status="재직",
            approval_status="승인완료",
            role=self.admin_role,
            must_change_password=False,
        )
        self.worker = Employee(
            name="보안 담당자",
            department=self.security,
            position="팀원",
            login_id="worker1",
            password_hash="",
            status="재직",
            approval_status="승인완료",
            role=self.admin_role,
            must_change_password=False,
        )
        self.admin.set_password("AdminPass123")
        self.worker.set_password("WorkerPass123")
        db.session.add_all([self.admin, self.worker])
        db.session.flush()
        self.network = WorkCategory(
            department=self.security,
            middle_name="정보화 기기",
            small_name="네트워크",
            active=True,
        )
        self.security_check = WorkCategory(
            department=self.security,
            middle_name="보안업무",
            small_name="보안점검",
            active=True,
        )
        self.accounting = WorkCategory(
            department=self.finance,
            middle_name="회계",
            small_name="월마감",
            active=True,
        )
        db.session.add_all([self.network, self.security_check, self.accounting])
        db.session.flush()
        self.task = Task(
            title="네트워크 구성 점검",
            content="사내 네트워크 상태 확인",
            department=self.security,
            work_category=self.network,
            assignee=self.worker,
            start_date=date(2026, 9, 4),
            target_date=date(2026, 9, 11),
            status="진행중",
            progress=10,
            repeat_cycle="주간",
            creator=self.admin,
        )
        self.task.classifications.append(TaskClassification(name="루틴"))
        db.session.add(self.task)
        db.session.commit()
        self.client = self.app.test_client()
        self.client.post(
            "/login",
            data={"login_id": "admin1", "password": "AdminPass123"},
        )

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.context.pop()

    def task_payload(self, category_id=None, department_id=None):
        return {
            "task_types": ["일반"],
            "title": "신규 네트워크 업무",
            "content": "네트워크 설정",
            "work_process": "1. 확인\n2. 적용",
            "department_id": str(department_id or self.security.id),
            "work_category_id": str(category_id or self.network.id),
            "assignee_id": str(self.worker.id),
            "start_date": "2026-09-04",
            "target_date": "2026-09-12",
            "status": "진행중",
            "progress": "0",
            "repeat_cycle": "없음",
            "repeat_detail": "",
        }

    def test_administrator_has_all_task_access_even_if_scope_is_own(self):
        dashboard = self.client.get("/")
        self.assertEqual(dashboard.status_code, 200)
        self.assertIn("보안전산팀".encode(), dashboard.data)
        page = self.client.get("/tasks")
        self.assertEqual(page.status_code, 200)
        self.assertIn(self.task.title.encode(), page.data)
        self.assertEqual(self.client.get(f"/tasks/{self.task.id}/edit").status_code, 200)
        self.assertIn("관리자는 모든 부서(팀)".encode(), page.data)

    def test_form_uses_department_middle_and_small_category_selectors(self):
        page = self.client.get("/tasks/new")
        self.assertEqual(page.status_code, 200)
        for expected in (
            "대분류 (부서(팀))",
            "중분류",
            "소분류",
            "data-work-category-form",
        ):
            self.assertIn(expected.encode(), page.data)
        self.assertIn(f'"id": {self.network.id}'.encode(), page.data)
        self.assertIn(b'"middle_name": "\\uc815\\ubcf4\\ud654 \\uae30\\uae30"', page.data)
        self.assertIn(b'"small_name": "\\ub124\\ud2b8\\uc6cc\\ud06c"', page.data)

    def test_selected_category_is_saved_and_shown_as_three_level_path(self):
        response = self.client.post("/tasks/new", data=self.task_payload())
        self.assertEqual(response.status_code, 302)
        task = db.session.scalar(
            db.select(Task).where(Task.title == "신규 네트워크 업무")
        )
        self.assertEqual(task.work_category_id, self.network.id)
        detail = self.client.get(f"/tasks/{task.id}")
        self.assertIn("보안전산팀".encode(), detail.data)
        self.assertIn("정보화 기기".encode(), detail.data)
        self.assertIn("네트워크".encode(), detail.data)

    def test_category_from_another_department_is_rejected(self):
        response = self.client.post(
            "/tasks/new",
            data=self.task_payload(category_id=self.accounting.id),
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("선택한 부서(팀)의 기초자료".encode(), response.data)
        self.assertIsNone(
            db.session.scalar(
                db.select(Task).where(Task.title == "신규 네트워크 업무")
            )
        )

    def test_list_can_filter_by_middle_and_small_category(self):
        matching = self.client.get(
            "/tasks?department_id={}&middle_category={}&small_category={}".format(
                self.security.id,
                "정보화+기기",
                "네트워크",
            )
        )
        self.assertIn(self.task.title.encode(), matching.data)
        excluded = self.client.get(
            f"/tasks?department_id={self.security.id}&middle_category=보안업무"
        )
        self.assertNotIn(self.task.title.encode(), excluded.data)

    def test_administrator_can_manage_category_reference_data(self):
        response = self.client.post(
            "/admin/work-categories",
            data={
                "department_id": str(self.security.id),
                "middle_name": "정보화 기기",
                "small_name": "서버",
            },
        )
        self.assertEqual(response.status_code, 302)
        category = db.session.scalar(
            db.select(WorkCategory).where(
                WorkCategory.department_id == self.security.id,
                WorkCategory.middle_name == "정보화 기기",
                WorkCategory.small_name == "서버",
            )
        )
        self.assertIsNotNone(category)
        page = self.client.get("/admin/work-categories")
        self.assertIn("업무구분 기초자료".encode(), page.data)
        self.assertIn("서버".encode(), page.data)

    def test_excel_template_adds_categories_and_legacy_template_still_uploads(self):
        response = self.client.get("/tasks/excel-template")
        workbook = load_workbook(io.BytesIO(response.data), data_only=True)
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers[3:6], ["부서(팀)", "중분류", "소분류"])

        legacy = Workbook()
        sheet = legacy.active
        sheet.append(
            [
                "제목",
                "내용",
                "분류",
                "부서",
                "담당자 로그인ID",
                "착수일",
                "목표일",
                "상태",
                "진행률",
                "반복주기",
            ]
        )
        sheet.append(
            [
                "기존 양식 업무",
                "기존 양식 호환",
                "일반",
                "보안전산팀",
                "worker1",
                date(2026, 9, 4),
                date(2026, 9, 5),
                "진행중",
                0,
                "없음",
            ]
        )
        upload = io.BytesIO()
        legacy.save(upload)
        upload.seek(0)
        result = self.client.post(
            "/tasks/excel-upload",
            data={"file": (upload, "legacy.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(result.status_code, 302)
        task = db.session.scalar(
            db.select(Task).where(Task.title == "기존 양식 업무")
        )
        self.assertIsNotNone(task)
        self.assertIsNone(task.work_category_id)

    def test_migration_backfills_excel_categories_without_changing_tasks(self):
        engine = sa.create_engine("sqlite://")
        metadata = sa.MetaData()
        departments = sa.Table(
            "departments",
            metadata,
            sa.Column("id", sa.Integer, primary_key=True),
            sa.Column("name", sa.String),
        )
        tasks = sa.Table(
            "tasks",
            metadata,
            sa.Column("id", sa.Integer, primary_key=True),
            sa.Column("department_id", sa.Integer, nullable=False),
            sa.Column("source_category", sa.String),
            sa.Column("source_detail", sa.String),
        )
        metadata.create_all(engine)
        with engine.begin() as connection:
            connection.execute(
                departments.insert(),
                {"id": 1, "name": "보안전산팀"},
            )
            connection.execute(
                tasks.insert(),
                [
                    {
                        "id": 1,
                        "department_id": 1,
                        "source_category": "정보화 기기",
                        "source_detail": "네트워크",
                    },
                    {
                        "id": 2,
                        "department_id": 1,
                        "source_category": "정보화 기기",
                        "source_detail": "네트워크",
                    },
                    {
                        "id": 3,
                        "department_id": 1,
                        "source_category": None,
                        "source_detail": None,
                    },
                ],
            )
            migration = importlib.import_module(
                "migrations.versions.20260904_0016_work_category_hierarchy"
            )
            original_op = migration.op
            migration.op = Operations(MigrationContext.configure(connection))
            try:
                migration.upgrade()
            finally:
                migration.op = original_op
            category_rows = connection.execute(
                sa.text(
                    "SELECT id, department_id, middle_name, small_name "
                    "FROM work_categories"
                )
            ).mappings().all()
            task_rows = connection.execute(
                sa.text("SELECT id, work_category_id FROM tasks ORDER BY id")
            ).mappings().all()

        self.assertEqual(len(category_rows), 1)
        self.assertEqual(category_rows[0]["middle_name"], "정보화 기기")
        self.assertEqual(category_rows[0]["small_name"], "네트워크")
        self.assertEqual(task_rows[0]["work_category_id"], category_rows[0]["id"])
        self.assertEqual(task_rows[1]["work_category_id"], category_rows[0]["id"])
        self.assertIsNone(task_rows[2]["work_category_id"])


if __name__ == "__main__":
    unittest.main()
