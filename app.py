import calendar
import hashlib
import io
import json
import os
import re
from datetime import date, datetime, timedelta, timezone
from functools import wraps
from urllib.parse import quote_plus, urlparse

from flask import (
    Flask,
    abort,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from flask_migrate import Migrate, upgrade
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook, load_workbook
from sqlalchemy import UniqueConstraint, and_, case, func, or_, select, text, update
from sqlalchemy.exc import IntegrityError
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash


db = SQLAlchemy()
migrate = Migrate(compare_type=True)
csrf = CSRFProtect()
login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.login_message = "로그인이 필요합니다."

TASK_TYPES = ("대표이사님 수명업무", "루틴", "일반", "주요")
CALENDAR_AUTO_TASK_TYPES = ("대표이사님 수명업무", "주요")
TASK_STATUSES = ("진행중", "완료", "지연", "보류")
REPEAT_CYCLES = ("없음", "일간", "주간", "월간", "분기", "반기", "연간", "상시")
STATUS_CLASS = {"진행중": "progress", "완료": "done", "지연": "delayed", "보류": "hold"}
WORK_CADENCES = (
    {"key": "daily", "label": "매일", "cycle": "일간", "aliases": ("일", "일간", "매일")},
    {"key": "weekly", "label": "매주", "cycle": "주간", "aliases": ("주", "주간", "매주", "주1회", "주2회")},
    {"key": "monthly", "label": "매월", "cycle": "월간", "aliases": ("월", "월간", "매월", "월1회", "월2회")},
    {"key": "quarterly", "label": "분기", "cycle": "분기", "aliases": ("분기", "분기별", "분기1회")},
    {"key": "semiannual", "label": "반기", "cycle": "반기", "aliases": ("반기", "반기별", "반기1회", "연2회", "년2회")},
    {"key": "yearly", "label": "매년", "cycle": "연간", "aliases": ("연", "년", "연간", "매년", "연1회", "년1회", "년1회이상")},
    {"key": "ongoing", "label": "상시", "cycle": "상시", "aliases": ("상시", "지속")},
    {"key": "irregular", "label": "비정기", "cycle": None, "aliases": ()},
)
WORK_CADENCE_BY_KEY = {item["key"]: item for item in WORK_CADENCES}
DASHBOARD_TASK_TYPES = (
    {"label": "대표이사님 수명업무", "type_name": "대표이사님 수명업무", "tone": "executive"},
    {"label": "루틴업무", "type_name": "루틴", "tone": "routine"},
    {"label": "일반업무", "type_name": "일반", "tone": "general"},
    {"label": "주요업무", "type_name": "주요", "tone": "major"},
)
DASHBOARD_DEPARTMENT_ORDER = ("재무운영팀", "인사총무팀", "보안전산팀")
EMPLOYEE_DEPARTMENTS = ("보안전산팀", "인사총무팀", "재무운영팀")
EMPLOYEE_POSITIONS = ("부서장", "팀장", "팀원")
MAX_FAILED_LOGIN_ATTEMPTS = 5
ACCOUNT_LOCK_MINUTES = 15
PASSWORD_MIN_LENGTH = 8
TEMP_PASSWORD_MIN_LENGTH = PASSWORD_MIN_LENGTH
REGISTRATION_PASSWORD_MIN_LENGTH = PASSWORD_MIN_LENGTH


def normalize_cadence_value(value):
    return str(value or "").replace(" ", "").strip()


def task_cadence_key(task):
    for cadence in WORK_CADENCES:
        if cadence["cycle"] and task.repeat_cycle == cadence["cycle"]:
            return cadence["key"]
    raw_values = {
        normalize_cadence_value(task.source_frequency),
        normalize_cadence_value(task.repeat_detail),
    }
    for cadence in WORK_CADENCES:
        if raw_values.intersection(cadence["aliases"]):
            return cadence["key"]
    return "irregular"


def task_cadence_condition(cadence_key):
    known_conditions = []
    normalized_frequency = func.replace(func.coalesce(Task.source_frequency, ""), " ", "")
    normalized_detail = func.replace(func.coalesce(Task.repeat_detail, ""), " ", "")
    for cadence in WORK_CADENCES:
        if cadence["key"] == "irregular":
            continue
        conditions = [Task.repeat_cycle == cadence["cycle"]]
        if cadence["aliases"]:
            conditions.extend(
                (
                    normalized_frequency.in_(cadence["aliases"]),
                    normalized_detail.in_(cadence["aliases"]),
                )
            )
        known_conditions.append((cadence["key"], or_(*conditions)))
    if cadence_key == "irregular":
        return ~or_(*(condition for _, condition in known_conditions))
    return dict(known_conditions)[cadence_key]


def utcnow():
    return datetime.now(timezone.utc)


def as_utc(value):
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def build_database_uri():
    if os.getenv("DATABASE_URL"):
        uri = os.environ["DATABASE_URL"]
        if uri.startswith("postgres://"):
            uri = uri.replace("postgres://", "postgresql+psycopg://", 1)
        elif uri.startswith("postgresql://"):
            uri = uri.replace("postgresql://", "postgresql+psycopg://", 1)
        return uri
    required = ["DB_HOST", "DB_PORT", "DB_NAME", "DB_USER", "DB_PASSWORD"]
    if all(os.getenv(key) for key in required):
        user = quote_plus(os.environ["DB_USER"])
        password = quote_plus(os.environ["DB_PASSWORD"])
        host = os.environ["DB_HOST"]
        port = os.environ["DB_PORT"]
        name = quote_plus(os.environ["DB_NAME"])
        return f"postgresql+psycopg://{user}:{password}@{host}:{port}/{name}"
    if os.getenv("ALLOW_IN_MEMORY_DB") == "1":
        return "sqlite:///:memory:"
    raise RuntimeError("PostgreSQL 연결 환경변수(DB_HOST/DB_PORT/DB_NAME/DB_USER/DB_PASSWORD)가 필요합니다.")


def build_secret_key():
    if os.getenv("SECRET_KEY"):
        return os.environ["SECRET_KEY"]
    database_secret = os.getenv("DB_PASSWORD")
    if database_secret:
        return hashlib.sha256(f"management-task:{database_secret}".encode()).hexdigest()
    if os.getenv("ALLOW_IN_MEMORY_DB") == "1":
        return "test-only-secret-key"
    raise RuntimeError("안전한 세션 키를 생성할 수 없습니다.")


role_menu = db.Table(
    "role_menu",
    db.Column("role_id", db.Integer, db.ForeignKey("roles.id", ondelete="CASCADE"), primary_key=True),
    db.Column("menu_id", db.Integer, db.ForeignKey("menus.id", ondelete="CASCADE"), primary_key=True),
)

role_board = db.Table(
    "role_board",
    db.Column("role_id", db.Integer, db.ForeignKey("roles.id", ondelete="CASCADE"), primary_key=True),
    db.Column("board_id", db.Integer, db.ForeignKey("boards.id", ondelete="CASCADE"), primary_key=True),
)

daily_meeting_item = db.Table(
    "daily_meeting_items",
    db.Column("meeting_id", db.Integer, db.ForeignKey("daily_meetings.id", ondelete="CASCADE"), primary_key=True),
    db.Column("task_id", db.Integer, db.ForeignKey("tasks.id", ondelete="CASCADE"), primary_key=True),
)


class Role(db.Model):
    __tablename__ = "roles"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    data_scope = db.Column(db.String(20), nullable=False, default="own")
    permissions = db.Column(db.JSON, nullable=False, default=dict)
    is_system = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    menus = db.relationship("Menu", secondary=role_menu, back_populates="roles")
    boards = db.relationship("Board", secondary=role_board, back_populates="roles")

    def allows(self, permission):
        return self.name == "관리자" or bool((self.permissions or {}).get(permission))


class Department(db.Model):
    __tablename__ = "departments"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=True)
    active = db.Column(db.Boolean, nullable=False, default=True)
    parent = db.relationship("Department", remote_side=[id], backref="children")

    @property
    def active_task_count(self):
        return sum(task.deleted_at is None for task in self.tasks)


class Employee(UserMixin, db.Model):
    __tablename__ = "employees"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    employee_no = db.Column(db.String(30), unique=True, nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=False)
    position = db.Column(db.String(50), nullable=True)
    email = db.Column(db.String(150), unique=True, nullable=True)
    phone = db.Column(db.String(30), nullable=True)
    login_id = db.Column(db.String(50), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(20), nullable=False, default="재직")
    hire_date = db.Column(db.Date, nullable=True)
    termination_date = db.Column(db.Date, nullable=True)
    reregistered_at = db.Column(db.DateTime(timezone=True), nullable=True)
    role_id = db.Column(db.Integer, db.ForeignKey("roles.id"), nullable=False)
    approval_status = db.Column(
        db.String(20), nullable=False, default="승인완료", server_default=text("'승인완료'")
    )
    approval_requested_at = db.Column(db.DateTime(timezone=True), nullable=True)
    approved_at = db.Column(db.DateTime(timezone=True), nullable=True)
    approved_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    failed_login_count = db.Column(db.Integer, nullable=False, default=0)
    locked_until = db.Column(db.DateTime(timezone=True), nullable=True)
    must_change_password = db.Column(db.Boolean, nullable=False, default=True)
    last_login_at = db.Column(db.DateTime(timezone=True), nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    department = db.relationship("Department", backref="employees")
    role = db.relationship("Role", backref="employees")
    approved_by = db.relationship("Employee", remote_side=[id], foreign_keys=[approved_by_id])

    @property
    def is_active(self):
        return self.status == "재직" and self.approval_status == "승인완료"

    def set_password(self, password):
        self.password_hash = generate_password_hash(password, method="scrypt")

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Menu(db.Model):
    __tablename__ = "menus"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    url = db.Column(db.String(200), nullable=False, default="#")
    parent_id = db.Column(db.Integer, db.ForeignKey("menus.id"), nullable=True)
    sort_order = db.Column(db.Integer, nullable=False, default=0)
    active = db.Column(db.Boolean, nullable=False, default=True)
    parent = db.relationship("Menu", remote_side=[id], backref="children")
    roles = db.relationship("Role", secondary=role_menu, back_populates="menus")


class Board(db.Model):
    __tablename__ = "boards"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    board_type = db.Column(db.String(30), nullable=False, default="일반")
    options = db.Column(db.JSON, nullable=False, default=dict)
    active = db.Column(db.Boolean, nullable=False, default=True)
    roles = db.relationship("Role", secondary=role_board, back_populates="boards")


class Task(db.Model):
    __tablename__ = "tasks"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(250), nullable=False)
    content = db.Column(db.Text, nullable=True)
    work_process = db.Column(db.Text, nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=False, index=True)
    assignee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False, index=True)
    start_date = db.Column(db.Date, nullable=False)
    target_date = db.Column(db.Date, nullable=False, index=True)
    completed_date = db.Column(db.Date, nullable=True)
    status = db.Column(db.String(20), nullable=False, default="진행중", index=True)
    progress = db.Column(db.Integer, nullable=False, default=0)
    repeat_cycle = db.Column(db.String(20), nullable=False, default="없음")
    repeat_detail = db.Column(db.String(100), nullable=True)
    calendar_selected = db.Column(db.Boolean, nullable=False, default=False, server_default=text("false"))
    calendar_excluded = db.Column(db.Boolean, nullable=False, default=False, server_default=text("false"))
    calendar_registered_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    source_ref = db.Column(db.String(100), unique=True, nullable=True, index=True)
    source_name = db.Column(db.String(255), nullable=True)
    source_sheet = db.Column(db.String(255), nullable=True)
    source_category = db.Column(db.String(100), nullable=True)
    source_detail = db.Column(db.String(150), nullable=True)
    source_content = db.Column(db.Text, nullable=True)
    source_assignees = db.Column(db.String(250), nullable=True)
    source_frequency = db.Column(db.String(50), nullable=True)
    created_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    updated_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow, onupdate=utcnow)
    deleted_at = db.Column(db.DateTime(timezone=True), nullable=True, index=True)
    deleted_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    department = db.relationship("Department", backref="tasks")
    assignee = db.relationship("Employee", foreign_keys=[assignee_id], backref="assigned_tasks")
    creator = db.relationship("Employee", foreign_keys=[created_by_id], backref="created_tasks")
    calendar_registered_by = db.relationship(
        "Employee", foreign_keys=[calendar_registered_by_id], backref="calendar_registered_tasks"
    )
    deleted_by = db.relationship("Employee", foreign_keys=[deleted_by_id])
    classifications = db.relationship("TaskClassification", cascade="all, delete-orphan", backref="task")
    daily_logs = db.relationship("TaskDailyLog", cascade="all, delete-orphan", backref="task")

    @property
    def type_names(self):
        return [item.name for item in self.classifications]

    @property
    def calendar_auto_included(self):
        return any(name in CALENDAR_AUTO_TASK_TYPES for name in self.type_names)

    @property
    def calendar_included(self):
        return not self.calendar_excluded and (self.calendar_auto_included or self.calendar_selected)

    @property
    def calendar_registration_label(self):
        if self.calendar_excluded:
            return "일정 제외"
        if self.calendar_auto_included:
            return "자동 등록"
        if self.calendar_selected:
            return "선택 등록"
        return "미등록"

    @property
    def remaining_days(self):
        if self.status == "완료":
            return None
        return (self.target_date - date.today()).days

    @property
    def status_class(self):
        return STATUS_CLASS.get(self.status, "progress")

    @property
    def display_assignees(self):
        return self.source_assignees or self.assignee.name

    @property
    def cadence_key(self):
        return task_cadence_key(self)

    @property
    def cadence_label(self):
        return WORK_CADENCE_BY_KEY[self.cadence_key]["label"]

    @property
    def cadence_detail(self):
        source_value = self.source_frequency or self.repeat_detail
        if source_value:
            return source_value
        if self.repeat_cycle != "없음":
            return self.repeat_cycle
        return None

    @property
    def is_source_import(self):
        return bool(self.source_ref)

    @property
    def dashboard_hidden(self):
        frequency = (self.source_frequency or self.repeat_detail or "").replace(" ", "")
        return self.repeat_cycle == "일간" or frequency in {"일", "일간", "매일", "지속", "상시"}

    @property
    def remaining_label(self):
        days = self.remaining_days
        if days is None:
            return "완료"
        if days == 0:
            return "D-DAY"
        if days < 0:
            return f"D+{abs(days)}"
        return f"D-{days}"


SOURCE_WORK_PROCESS_IDENTITY_FIELDS = (
    "source_name",
    "source_sheet",
    "source_category",
    "source_detail",
    "source_content",
)


def source_work_process_identity(task):
    """Return the immutable source identity used to share a work process."""
    if not task.source_ref:
        return None
    return tuple(getattr(task, field) for field in SOURCE_WORK_PROCESS_IDENTITY_FIELDS)


def source_work_process_condition(task):
    identity = source_work_process_identity(task)
    if identity is None:
        return None
    conditions = [Task.source_ref.is_not(None), Task.deleted_at.is_(None)]
    for field, value in zip(SOURCE_WORK_PROCESS_IDENTITY_FIELDS, identity):
        column = getattr(Task, field)
        conditions.append(column.is_(None) if value is None else column == value)
    return and_(*conditions)


def synchronize_task_work_process(task):
    """Apply one imported task's process to every exact source match."""
    condition = source_work_process_condition(task)
    if condition is None:
        return []
    matching_tasks = db.session.scalars(
        select(Task).where(condition, Task.id != task.id)
    ).all()
    changed_ids = []
    for matching_task in matching_tasks:
        if matching_task.work_process != task.work_process:
            matching_task.work_process = task.work_process
            changed_ids.append(matching_task.id)
    return changed_ids


def synchronize_existing_source_work_processes():
    """Backfill exact source groups from their most recently updated process."""
    source_tasks = db.session.scalars(
        select(Task)
        .where(Task.source_ref.is_not(None), Task.deleted_at.is_(None))
        .order_by(Task.updated_at.desc(), Task.id.desc())
    ).all()
    groups = {}
    for task in source_tasks:
        groups.setdefault(source_work_process_identity(task), []).append(task)

    changed_ids = []
    for tasks in groups.values():
        latest_process = next((task.work_process for task in tasks if task.work_process), None)
        if not latest_process:
            continue
        for task in tasks:
            if task.work_process != latest_process:
                task.work_process = latest_process
                changed_ids.append(task.id)
    return changed_ids


class TaskClassification(db.Model):
    __tablename__ = "task_classifications"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False)
    name = db.Column(db.String(30), nullable=False)
    __table_args__ = (UniqueConstraint("task_id", "name", name="uq_task_classification"),)


class TaskStatusLog(db.Model):
    __tablename__ = "task_status_logs"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False, index=True)
    previous_status = db.Column(db.String(20), nullable=True)
    new_status = db.Column(db.String(20), nullable=False)
    changed_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    changed_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    changer = db.relationship("Employee")


class TaskDailyLog(db.Model):
    __tablename__ = "task_daily_logs"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False, index=True)
    work_date = db.Column(db.Date, nullable=False, default=date.today, index=True)
    content = db.Column(db.Text, nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    author = db.relationship("Employee")


class WorkJournalItem(db.Model):
    __tablename__ = "work_journal_items"
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id", ondelete="CASCADE"), nullable=False, index=True)
    work_date = db.Column(db.Date, nullable=False, default=date.today, index=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False, index=True)
    added_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    task = db.relationship("Task", backref=db.backref("work_journal_items", cascade="all, delete-orphan"))
    employee = db.relationship("Employee", foreign_keys=[employee_id])
    added_by = db.relationship("Employee", foreign_keys=[added_by_id])
    __table_args__ = (
        UniqueConstraint("task_id", "work_date", "employee_id", name="uq_work_journal_item"),
    )


class Schedule(db.Model):
    __tablename__ = "schedules"
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    schedule_date = db.Column(db.Date, nullable=False, index=True)
    task_id = db.Column(db.Integer, db.ForeignKey("tasks.id", ondelete="SET NULL"), nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=True)
    assignee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    scope = db.Column(db.String(20), nullable=False, default="개인")
    memo = db.Column(db.Text, nullable=True)
    is_holiday = db.Column(db.Boolean, nullable=False, default=False)
    created_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    deleted_at = db.Column(db.DateTime(timezone=True), nullable=True, index=True)
    deleted_by_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    task = db.relationship("Task")
    department = db.relationship("Department")
    assignee = db.relationship("Employee", foreign_keys=[assignee_id])
    creator = db.relationship("Employee", foreign_keys=[created_by_id])
    deleted_by = db.relationship("Employee", foreign_keys=[deleted_by_id])


class DailyMeeting(db.Model):
    __tablename__ = "daily_meetings"
    id = db.Column(db.Integer, primary_key=True)
    meeting_date = db.Column(db.Date, nullable=False, index=True)
    author_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey("departments.id"), nullable=False)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    author = db.relationship("Employee")
    department = db.relationship("Department")
    tasks = db.relationship("Task", secondary=daily_meeting_item)


class Attachment(db.Model):
    __tablename__ = "attachments"
    id = db.Column(db.Integer, primary_key=True)
    related_type = db.Column(db.String(30), nullable=False)
    related_id = db.Column(db.Integer, nullable=True)
    filename = db.Column(db.String(255), nullable=False)
    content_type = db.Column(db.String(120), nullable=True)
    file_data = db.Column(db.LargeBinary, nullable=False)
    uploader_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    uploaded_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow)
    uploader = db.relationship("Employee")


class AuditLog(db.Model):
    __tablename__ = "audit_logs"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True, index=True)
    action = db.Column(db.String(80), nullable=False)
    target = db.Column(db.String(150), nullable=False)
    details = db.Column(db.JSON, nullable=False, default=dict)
    ip_address = db.Column(db.String(64), nullable=True)
    created_at = db.Column(db.DateTime(timezone=True), nullable=False, default=utcnow, index=True)
    user = db.relationship("Employee")


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Employee, int(user_id))


def audit(action, target, details=None, user_id=None):
    uid = user_id if user_id is not None else (current_user.id if current_user.is_authenticated else None)
    db.session.add(
        AuditLog(
            user_id=uid,
            action=action,
            target=target,
            details=details or {},
            ip_address=request.headers.get("X-Forwarded-For", request.remote_addr) if request else None,
        )
    )


def admin_required(permission):
    def decorator(fn):
        @wraps(fn)
        @login_required
        def wrapped(*args, **kwargs):
            if not current_user.role.allows(permission):
                abort(403)
            return fn(*args, **kwargs)

        return wrapped

    return decorator


def visible_task_query(user):
    query = select(Task).where(Task.deleted_at.is_(None))
    scope = user.role.data_scope
    if scope == "all":
        return query
    if scope == "department":
        return query.where(Task.department_id == user.department_id)
    return query.where(or_(Task.assignee_id == user.id, Task.created_by_id == user.id))


def can_view_task_detail(task):
    if task.deleted_at is not None:
        return False
    return (
        current_user.role.data_scope == "all"
        or task.department_id == current_user.department_id
        or task.assignee_id == current_user.id
        or task.created_by_id == current_user.id
    )


def can_edit_task(task):
    if task.deleted_at is not None:
        return False
    return current_user.role.allows("task_manage_all") or (
        current_user.role.allows("task_manage_department")
        and task.department_id == current_user.department_id
    )


def can_delete_task(task):
    if task.deleted_at is not None:
        return False
    if current_user.role.name == "관리자":
        return True
    return current_user.role.name in {"부서장", "팀장"} and task.department_id == current_user.department_id


def can_remove_task_calendar(task):
    if task.deleted_at is not None or not task.calendar_included:
        return False
    if current_user.role.name == "관리자":
        return True
    if current_user.role.name in {"부서장", "팀장"}:
        return task.department_id == current_user.department_id
    if current_user.role.name != "팀원" or task.department_id != current_user.department_id:
        return False
    if task.calendar_registered_by_id is not None:
        return task.calendar_registered_by_id == current_user.id
    return task.created_by_id == current_user.id


def get_active_task_or_404(task_id):
    task = db.session.scalar(
        select(Task).where(Task.id == task_id, Task.deleted_at.is_(None))
    )
    if not task:
        abort(404)
    return task


def can_write_department(department_id):
    return current_user.role.allows("task_manage_all") or department_id == current_user.department_id


def can_view_all_schedules(user):
    return user.role.name in {"관리자", "부서장", "팀장"} or user.role.allows("schedule_view_all")


def visible_schedule_query(user):
    query = select(Schedule).where(Schedule.deleted_at.is_(None))
    if can_view_all_schedules(user):
        return query
    department_employee_ids = select(Employee.id).where(
        Employee.department_id == user.department_id,
        Employee.status == "재직",
        Employee.approval_status == "승인완료",
    )
    return query.where(
        or_(
            Schedule.scope == "전사",
            Schedule.department_id == user.department_id,
            Schedule.assignee_id.in_(department_employee_ids),
        )
    )


def can_delete_schedule(schedule):
    if schedule.deleted_at is not None:
        return False
    if current_user.role.name == "관리자" or schedule.created_by_id == current_user.id:
        return True
    return (
        current_user.role.name in {"부서장", "팀장"}
        and schedule.department_id == current_user.department_id
    )


def summarize_tasks(tasks, type_name):
    subset = [task for task in tasks if type_name in task.type_names]
    return {
        "total": len(subset),
        "today": sum(1 for task in subset if task.start_date <= date.today() <= task.target_date),
        "ongoing": sum(1 for task in subset if task.status == "진행중"),
        "done": sum(1 for task in subset if task.status == "완료"),
        "delayed": sum(1 for task in subset if task.status == "지연"),
        "hold": sum(1 for task in subset if task.status == "보류"),
    }


def build_department_summaries(departments, tasks):
    summaries = []
    for department in departments:
        department_tasks = [task for task in tasks if task.department_id == department.id]
        summaries.append(
            {
                "department": department,
                "total": len(department_tasks),
                "routine": summarize_tasks(department_tasks, "루틴"),
                "major": summarize_tasks(department_tasks, "주요"),
                "executive": summarize_tasks(department_tasks, "대표이사님 수명업무"),
                "general": summarize_tasks(department_tasks, "일반"),
                "ongoing": sum(1 for task in department_tasks if task.status == "진행중"),
                "delayed": sum(1 for task in department_tasks if task.status == "지연"),
                "hold": sum(1 for task in department_tasks if task.status == "보류"),
            }
        )
    return summaries


def build_task_type_summaries(tasks, departments):
    summaries = []
    for spec in DASHBOARD_TASK_TYPES:
        type_tasks = [task for task in tasks if spec["type_name"] in task.type_names]
        status_summary = summarize_tasks(tasks, spec["type_name"])
        summaries.append(
            {
                **spec,
                "total": status_summary["total"],
                "department_counts": [
                    {
                        "department": department,
                        "count": sum(1 for task in type_tasks if task.department_id == department.id),
                    }
                    for department in departments
                ],
                "ongoing": status_summary["ongoing"],
                "delayed": status_summary["delayed"],
                "hold": status_summary["hold"],
            }
        )
    return summaries


def build_cadence_summaries(tasks):
    return [
        {
            **cadence,
            "count": sum(1 for task in tasks if task.cadence_key == cadence["key"]),
        }
        for cadence in WORK_CADENCES
    ]


def refresh_overdue_tasks():
    overdue = db.session.scalars(
        select(Task).where(
            Task.deleted_at.is_(None),
            Task.target_date < date.today(),
            Task.status == "진행중",
        )
    ).all()
    if not overdue:
        return
    for task in overdue:
        task.status = "지연"
        db.session.add(
            TaskStatusLog(task_id=task.id, previous_status="진행중", new_status="지연", changed_by_id=None)
        )
    db.session.commit()


def import_security_it_tasks():
    seed_path = os.path.join(os.path.dirname(__file__), "seed", "security_it_tasks_260902.json")
    with open(seed_path, encoding="utf-8") as seed_file:
        payload = json.load(seed_file)

    department = db.session.scalar(select(Department).where(Department.name == "보안전산팀"))
    admin_user = db.session.scalar(
        select(Employee).where(Employee.login_id == os.getenv("BOOTSTRAP_ADMIN_ID", "admin"))
    )
    if not department or not admin_user:
        raise RuntimeError("보안전산팀 업무를 등록할 부서 또는 관리자 계정을 찾을 수 없습니다.")

    recurring_frequencies = {"일", "주1회", "월1회", "월2회", "분기별", "년1회", "년1회이상", "지속"}
    repeat_cycles = {
        "일": "일간",
        "주1회": "주간",
        "월1회": "월간",
        "월2회": "월간",
        "분기별": "분기",
        "년1회": "연간",
        "년1회이상": "연간",
    }
    target_dates = {
        "일": date(2026, 9, 2),
        "주1회": date(2026, 9, 9),
        "월1회": date(2026, 9, 30),
        "월2회": date(2026, 9, 30),
        "분기별": date(2026, 9, 30),
        "년1회": date(2026, 12, 31),
        "년1회이상": date(2026, 12, 31),
        "지속": date(2026, 12, 31),
        "변경 시": date(2026, 12, 31),
        "불시": date(2026, 12, 31),
        "단발성": date(2026, 12, 31),
    }
    existing_refs = set(
        db.session.scalars(
            select(Task.source_ref).where(Task.source_ref.in_([item["source_ref"] for item in payload["tasks"]]))
        ).all()
    )
    created_count = 0
    for item in payload["tasks"]:
        if item["source_ref"] in existing_refs:
            continue
        frequency = item["frequency"]
        title_detail = item["detail"] or item["content"]
        task = Task(
            title=f'{item["category"]} · {title_detail}'[:250],
            content=item["content"],
            department_id=department.id,
            assignee_id=admin_user.id,
            start_date=date(2026, 9, 2),
            target_date=target_dates.get(frequency, date(2026, 12, 31)),
            status="진행중",
            progress=0,
            repeat_cycle=repeat_cycles.get(frequency, "없음"),
            repeat_detail=frequency,
            source_ref=item["source_ref"],
            source_name=item["source_name"],
            source_sheet=item["source_sheet"],
            source_category=item["category"],
            source_detail=item["detail"],
            source_content=item["content"],
            source_assignees=", ".join(item["assignees"]),
            source_frequency=frequency,
            created_by_id=admin_user.id,
        )
        classification = "루틴" if frequency in recurring_frequencies else "주요"
        task.classifications = [TaskClassification(name=classification)]
        db.session.add(task)
        created_count += 1

    imported_count = db.session.scalar(
        select(func.count(Task.id)).where(Task.source_name == "경영_보안전산팀 업무분장_260902.xlsx")
    ) or 0
    final_count = imported_count
    if final_count != payload["expected_count"]:
        db.session.rollback()
        raise RuntimeError(
            f'보안전산팀 업무 등록 건수가 일치하지 않습니다: {final_count}/{payload["expected_count"]}'
        )
    if created_count:
        db.session.add(
            AuditLog(
                user_id=admin_user.id,
                action="SECURITY_IT_WORKBOOK_IMPORT",
                target="경영_보안전산팀 업무분장_260902.xlsx",
                details={"created": created_count, "total": final_count, "deduplicated": 8},
                ip_address=None,
            )
        )
    db.session.commit()


def seed_reference_data():
    for legacy_name, current_name in (("대표이사수명", "대표이사님 수명업무"), ("개인", "일반")):
        legacy_task_types = db.session.scalars(
            select(TaskClassification).where(TaskClassification.name == legacy_name)
        ).all()
        for legacy_type in legacy_task_types:
            current_type = db.session.scalar(
                select(TaskClassification).where(
                    TaskClassification.task_id == legacy_type.task_id,
                    TaskClassification.name == current_name,
                )
            )
            if current_type:
                db.session.delete(legacy_type)
            else:
                legacy_type.name = current_name
    db.session.flush()

    for legacy_name, current_name in (("시스템관리자", "관리자"), ("대표이사", "부서장")):
        legacy_role = db.session.scalar(select(Role).where(Role.name == legacy_name))
        current_role = db.session.scalar(select(Role).where(Role.name == current_name))
        if legacy_role and not current_role:
            legacy_role.name = current_name
        elif legacy_role and current_role:
            for employee in list(legacy_role.employees):
                employee.role = current_role
            for menu in list(legacy_role.menus):
                if current_role not in menu.roles:
                    menu.roles.append(current_role)
            for board in list(legacy_role.boards):
                if current_role not in board.roles:
                    board.roles.append(current_role)
            db.session.delete(legacy_role)
    db.session.flush()

    role_specs = [
        ("관리자", "all", {"admin": True, "task_create": True, "task_manage_all": True, "meeting_all": True, "schedule_view_all": True}, True),
        ("부서장", "all", {"task_create": True, "task_view_all": True, "task_manage_department": True, "meeting_all": True, "schedule_view_all": True}, True),
        ("팀장", "all", {"task_create": True, "task_view_all": True, "task_manage_department": True, "meeting_all": True, "schedule_view_all": True}, True),
        ("팀원", "department", {"task_create": True, "task_manage_department": True}, True),
    ]
    roles = {}
    for name, scope, permissions, is_system in role_specs:
        role = db.session.scalar(select(Role).where(Role.name == name))
        if not role:
            role = Role(name=name, data_scope=scope, permissions=permissions, is_system=is_system)
            db.session.add(role)
        else:
            role.data_scope = scope
            role.permissions = permissions
            role.is_system = is_system
        roles[name] = role

    legacy_task_menu = db.session.scalar(select(Menu).where(Menu.name == "업무현황"))
    current_task_menu = db.session.scalar(select(Menu).where(Menu.name == "각 부서 업무 현황"))
    if legacy_task_menu and not current_task_menu:
        legacy_task_menu.name = "각 부서 업무 현황"
    elif legacy_task_menu and current_task_menu:
        for role in list(legacy_task_menu.roles):
            if role not in current_task_menu.roles:
                current_task_menu.roles.append(role)
        db.session.delete(legacy_task_menu)
    db.session.flush()

    legacy_schedule_menu = db.session.scalar(
        select(Menu).where(Menu.name == "일정", Menu.url == "/calendar")
    )
    current_schedule_menu = db.session.scalar(select(Menu).where(Menu.name == "일정(캘린더)"))
    if legacy_schedule_menu and not current_schedule_menu:
        legacy_schedule_menu.name = "일정(캘린더)"
    elif legacy_schedule_menu and current_schedule_menu:
        for role in list(legacy_schedule_menu.roles):
            if role not in current_schedule_menu.roles:
                current_schedule_menu.roles.append(role)
        db.session.delete(legacy_schedule_menu)
    db.session.flush()

    for name in EMPLOYEE_DEPARTMENTS:
        department = db.session.scalar(select(Department).where(Department.name == name))
        if not department:
            department = Department(name=name)
            db.session.add(department)
        department.active = True
        department.parent_id = None
    db.session.flush()

    for division_name in ("경영사업본부", "마케팅사업본부", "기술사업본부"):
        division = db.session.scalar(select(Department).where(Department.name == division_name))
        if not division:
            continue
        for child in list(division.children):
            child.parent_id = None
        db.session.execute(
            update(Schedule).where(Schedule.department_id == division.id).values(department_id=None)
        )
        db.session.flush()
        hard_reference_count = sum(
            db.session.scalar(select(func.count(model.id)).where(model.department_id == division.id)) or 0
            for model in (Employee, Task, DailyMeeting)
        )
        if hard_reference_count:
            division.active = False
        else:
            db.session.delete(division)
        db.session.flush()

    menu_specs = [
        ("통합현황", "/", 10),
        ("각 부서 업무 현황", "/tasks", 20),
        ("업무등록", "/tasks/new", 30),
        ("일정(캘린더)", "/calendar", 40),
        ("일일회의", "/meetings", 50),
        ("업무일지", "/journals", 60),
        ("관리자", "/admin/employees", 90),
    ]
    for name, url, order in menu_specs:
        menu = db.session.scalar(select(Menu).where(Menu.name == name))
        if not menu:
            menu = Menu(name=name, url=url, sort_order=order)
            db.session.add(menu)
        menu.roles = list(roles.values()) if name != "관리자" else [roles["관리자"]]
    board = db.session.scalar(select(Board).where(Board.name == "공지사항"))
    if not board:
        board = Board(name="공지사항", board_type="공지", options={"allow_comments": False})
        db.session.add(board)
    board.roles = list(roles.values())

    bootstrap_id = os.getenv("BOOTSTRAP_ADMIN_ID", "admin")
    bootstrap_hash = os.getenv("BOOTSTRAP_ADMIN_PASSWORD_HASH")
    if bootstrap_hash and not db.session.scalar(select(Employee).where(Employee.login_id == bootstrap_id)):
        dept = db.session.scalar(select(Department).where(Department.name == "보안전산팀"))
        if not dept:
            raise RuntimeError("관리자 계정용 보안전산팀 부서를 찾을 수 없습니다.")
        db.session.add(
            Employee(
                name=os.getenv("BOOTSTRAP_ADMIN_NAME", "시스템관리자"),
                employee_no="SYSTEM-001",
                department=dept,
                position="관리자",
                login_id=bootstrap_id,
                password_hash=bootstrap_hash,
                role=roles["관리자"],
                status="재직",
                hire_date=date.today(),
                must_change_password=True,
            )
        )

    password_reset_hash = os.getenv("ADMIN_PASSWORD_RESET_HASH")
    if password_reset_hash:
        admin_user = db.session.scalar(select(Employee).where(Employee.login_id == bootstrap_id))
        if not admin_user:
            raise RuntimeError(f"비밀번호를 변경할 관리자 계정({bootstrap_id})을 찾을 수 없습니다.")
        if admin_user.password_hash != password_reset_hash:
            admin_user.password_hash = password_reset_hash
            admin_user.failed_login_count = 0
            admin_user.locked_until = None
            admin_user.must_change_password = False
            db.session.add(
                AuditLog(
                    user_id=admin_user.id,
                    action="ADMIN_PASSWORD_RESET",
                    target=f"employee:{admin_user.id}",
                    details={"source": "one_time_environment"},
                    ip_address=None,
                )
            )
    db.session.commit()
    import_security_it_tasks()
    if synchronize_existing_source_work_processes():
        db.session.commit()


def run_startup_tasks(app):
    if app.config.get("TESTING"):
        return
    with app.app_context():
        uri = app.config["SQLALCHEMY_DATABASE_URI"]
        try:
            if uri.startswith("postgresql"):
                db.session.execute(text("SELECT pg_advisory_lock(2026090201)"))
                db.session.commit()
            upgrade(directory=os.path.join(app.root_path, "migrations"))
            seed_reference_data()
        finally:
            if uri.startswith("postgresql"):
                try:
                    db.session.execute(text("SELECT pg_advisory_unlock(2026090201)"))
                    db.session.commit()
                except Exception:
                    db.session.rollback()


def create_app(test_config=None):
    app = Flask(__name__)
    app.config.update(
        SECRET_KEY=build_secret_key(),
        SQLALCHEMY_DATABASE_URI=build_database_uri(),
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        SQLALCHEMY_ENGINE_OPTIONS={"pool_pre_ping": True, "pool_recycle": 280},
        MAX_CONTENT_LENGTH=10 * 1024 * 1024,
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
        SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE", "1") == "1",
        PERMANENT_SESSION_LIFETIME=timedelta(minutes=int(os.getenv("SESSION_TIMEOUT_MINUTES", "60"))),
    )
    if test_config:
        app.config.update(test_config)
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
    db.init_app(app)
    migrate.init_app(app, db, directory=os.path.join(app.root_path, "migrations"))
    csrf.init_app(app)
    login_manager.init_app(app)

    @app.before_request
    def enforce_authentication():
        public_endpoints = {"login", "employee_registration", "health", "static"}
        if request.endpoint not in public_endpoints and not current_user.is_authenticated:
            return redirect(url_for("login", next=request.full_path))
        if current_user.is_authenticated:
            if not current_user.is_active:
                logout_user()
                return redirect(url_for("login"))
            if current_user.must_change_password and request.endpoint not in {"change_password", "logout", "static"}:
                return redirect(url_for("change_password"))
            refresh_overdue_tasks()

    @app.context_processor
    def inject_globals():
        menus = []
        if current_user.is_authenticated:
            menus = db.session.scalars(
                select(Menu)
                .join(role_menu)
                .where(role_menu.c.role_id == current_user.role_id, Menu.active.is_(True))
                .order_by(Menu.sort_order)
            ).all()
        return {
            "visible_menus": menus,
            "TASK_TYPES": TASK_TYPES,
            "TASK_STATUSES": TASK_STATUSES,
            "REPEAT_CYCLES": REPEAT_CYCLES,
            "STATUS_CLASS": STATUS_CLASS,
            "PASSWORD_MIN_LENGTH": PASSWORD_MIN_LENGTH,
            "REGISTRATION_PASSWORD_MIN_LENGTH": REGISTRATION_PASSWORD_MIN_LENGTH,
            "today": date.today(),
        }

    @app.get("/health")
    def health():
        try:
            db.session.execute(text("SELECT 1"))
            admin_ready = bool(
                db.session.scalar(
                    select(func.count(Employee.id))
                    .join(Role, Employee.role_id == Role.id)
                    .where(
                        Role.name == "관리자",
                        Employee.status == "재직",
                        Employee.approval_status == "승인완료",
                    )
                )
            )
            reset_hash = os.getenv("ADMIN_PASSWORD_RESET_HASH")
            reset_ready = True
            if reset_hash:
                reset_user = db.session.scalar(
                    select(Employee).where(Employee.login_id == os.getenv("BOOTSTRAP_ADMIN_ID", "admin"))
                )
                reset_ready = bool(reset_user and reset_user.password_hash == reset_hash)
            import_ready = (
                db.session.scalar(
                    select(func.count(Task.id)).where(
                        Task.source_name == "경영_보안전산팀 업무분장_260902.xlsx"
                    )
                )
                == 116
            )
            ready = admin_ready and reset_ready and import_ready
            return jsonify(
                status="ok",
                database="connected",
                application_ready=admin_ready,
                credential_sync_ready=reset_ready,
                security_it_import_ready=import_ready,
            ), (200 if ready else 503)
        except Exception:
            return jsonify(status="error", database="unavailable"), 503

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            login_id = request.form.get("login_id", "").strip()
            password = request.form.get("password", "")
            user = db.session.scalar(select(Employee).where(Employee.login_id == login_id))
            now = utcnow()
            locked_until = as_utc(user.locked_until) if user else None
            if locked_until and locked_until > now:
                flash("로그인 실패 횟수 초과로 계정이 잠겼습니다. 잠시 후 다시 시도해 주세요.", "danger")
                return render_template("auth.html", mode="login"), 429
            if not user or not user.check_password(password):
                if user:
                    user.failed_login_count += 1
                    if user.failed_login_count >= MAX_FAILED_LOGIN_ATTEMPTS:
                        user.locked_until = now + timedelta(minutes=ACCOUNT_LOCK_MINUTES)
                        user.failed_login_count = 0
                    db.session.commit()
                flash("아이디 또는 비밀번호를 확인해 주세요.", "danger")
                return render_template("auth.html", mode="login"), 401
            if user.approval_status == "승인대기":
                flash("임직원 계정 등록 신청이 관리자 승인 대기 중입니다.", "danger")
                return render_template("auth.html", mode="login"), 403
            if not user.is_active:
                flash("사용할 수 없는 계정입니다. 관리자에게 문의해 주세요.", "danger")
                return render_template("auth.html", mode="login"), 403
            user.failed_login_count = 0
            user.locked_until = None
            user.last_login_at = now
            db.session.commit()
            login_user(user, remember=False, duration=app.config["PERMANENT_SESSION_LIFETIME"])
            audit("LOGIN", f"employee:{user.id}", user_id=user.id)
            db.session.commit()
            next_url = request.args.get("next", "")
            if next_url and urlparse(next_url).netloc == "":
                return redirect(next_url)
            return redirect(url_for("dashboard"))
        return render_template("auth.html", mode="login")

    @app.route("/register", methods=["GET", "POST"])
    def employee_registration():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        departments = db.session.scalars(
            select(Department).where(
                Department.active.is_(True),
                Department.name.in_(EMPLOYEE_DEPARTMENTS),
            )
        ).all()
        departments_by_name = {department.name: department for department in departments}
        departments = [
            departments_by_name[name] for name in EMPLOYEE_DEPARTMENTS if name in departments_by_name
        ]
        if request.method == "POST":
            try:
                login_id = request.form.get("login_id", "").strip()
                if not re.fullmatch(r"[A-Za-z0-9]{4,30}", login_id):
                    raise ValueError("계정 ID는 아마란스 계정의 영문·숫자 4~30자로 입력해 주세요.")
                name = request.form.get("name", "").strip()
                if not name:
                    raise ValueError("성명을 입력해 주세요.")
                department_id = request.form.get("department_id", type=int)
                department = db.session.get(Department, department_id) if department_id else None
                if not department or not department.active or department.name not in EMPLOYEE_DEPARTMENTS:
                    raise ValueError("부서(팀)를 선택해 주세요.")
                position = request.form.get("position", "").strip()
                if position not in EMPLOYEE_POSITIONS:
                    raise ValueError("직급을 선택해 주세요.")
                password = request.form.get("password", "")
                if len(password) < REGISTRATION_PASSWORD_MIN_LENGTH or not any(character.isalpha() for character in password) or not any(
                    character.isdigit() for character in password
                ):
                    raise ValueError(
                        f"비밀번호는 영문과 숫자를 포함하여 {REGISTRATION_PASSWORD_MIN_LENGTH}자 이상이어야 합니다."
                    )
                if password != request.form.get("confirm_password", ""):
                    raise ValueError("비밀번호 확인이 일치하지 않습니다.")
                default_role = db.session.scalar(select(Role).where(Role.name == "팀원"))
                if not default_role:
                    raise ValueError("기본 권한을 확인할 수 없습니다. 관리자에게 문의해 주세요.")
                employee = Employee(
                    name=name,
                    employee_no=request.form.get("employee_no", "").strip() or None,
                    department_id=department.id,
                    position=position,
                    email=request.form.get("email", "").strip() or None,
                    phone=request.form.get("phone", "").strip() or None,
                    login_id=login_id,
                    role_id=default_role.id,
                    status="재직",
                    hire_date=(
                        date.fromisoformat(request.form["hire_date"])
                        if request.form.get("hire_date")
                        else None
                    ),
                    password_hash="",
                    must_change_password=False,
                    approval_status="승인대기",
                    approval_requested_at=utcnow(),
                )
                employee.set_password(password)
                db.session.add(employee)
                db.session.flush()
                audit(
                    "EMPLOYEE_REGISTRATION_REQUEST",
                    f"employee:{employee.id}",
                    {
                        "login_id": employee.login_id,
                        "department_id": employee.department_id,
                        "default_role": default_role.name,
                        "hire_date_provided": employee.hire_date is not None,
                    },
                    user_id=None,
                )
                db.session.commit()
                flash("임직원 계정 등록 신청이 완료되었습니다. 관리자 승인 후 로그인할 수 있습니다.", "success")
                return redirect(url_for("login"))
            except IntegrityError:
                db.session.rollback()
                flash("계정 ID, 사번 또는 이메일이 이미 등록되어 있습니다.", "danger")
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            return render_template(
                "auth.html",
                mode="register",
                employee_departments=departments,
                employee_positions=EMPLOYEE_POSITIONS,
            ), 400
        return render_template(
            "auth.html",
            mode="register",
            employee_departments=departments,
            employee_positions=EMPLOYEE_POSITIONS,
        )

    @app.post("/logout")
    @login_required
    def logout():
        audit("LOGOUT", f"employee:{current_user.id}")
        db.session.commit()
        logout_user()
        return redirect(url_for("login"))

    @app.route("/change-password", methods=["GET", "POST"])
    @login_required
    def change_password():
        if request.method == "POST":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not current_user.check_password(current_password):
                flash("현재 비밀번호가 일치하지 않습니다.", "danger")
            elif len(new_password) < PASSWORD_MIN_LENGTH or not any(c.isalpha() for c in new_password) or not any(c.isdigit() for c in new_password):
                flash(
                    f"새 비밀번호는 영문과 숫자를 포함하여 {PASSWORD_MIN_LENGTH}자 이상이어야 합니다.",
                    "danger",
                )
            elif new_password != confirm_password:
                flash("새 비밀번호 확인이 일치하지 않습니다.", "danger")
            else:
                current_user.set_password(new_password)
                current_user.must_change_password = False
                audit("PASSWORD_CHANGE", f"employee:{current_user.id}")
                db.session.commit()
                flash("비밀번호가 변경되었습니다.", "success")
                return redirect(url_for("dashboard"))
        return render_template("auth.html", mode="change")

    @app.route("/profile", methods=["GET", "POST"])
    @login_required
    def profile():
        if request.method == "POST":
            try:
                name = request.form.get("name", "").strip()
                if not name:
                    raise ValueError("성명을 입력해 주세요.")
                current_user.name = name
                current_user.employee_no = request.form.get("employee_no", "").strip() or None
                current_user.email = request.form.get("email", "").strip() or None
                current_user.phone = request.form.get("phone", "").strip() or None
                audit(
                    "EMPLOYEE_SELF_UPDATE",
                    f"employee:{current_user.id}",
                    {"fields": ["name", "employee_no", "email", "phone"]},
                )
                db.session.commit()
                flash("내 정보가 수정되었습니다.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("사번 또는 이메일이 이미 등록되어 있습니다.", "danger")
            except ValueError as exc:
                db.session.rollback()
                flash(f"저장할 수 없습니다: {exc}", "danger")
            return redirect(url_for("profile"))
        return render_template("profile.html")

    @app.get("/")
    @login_required
    def dashboard():
        tasks = db.session.scalars(visible_task_query(current_user).order_by(Task.target_date, Task.id)).all()

        departments_query = select(Department).where(Department.active.is_(True))
        if current_user.role.data_scope != "all":
            departments_query = departments_query.where(Department.id == current_user.department_id)
        visible_departments = db.session.scalars(departments_query).all()
        dashboard_department_order = {
            department_name: index for index, department_name in enumerate(DASHBOARD_DEPARTMENT_ORDER)
        }
        visible_departments.sort(
            key=lambda department: (
                dashboard_department_order.get(department.name, len(dashboard_department_order)),
                department.name,
            )
        )
        board_tasks = [task for task in tasks if not task.dashboard_hidden]
        routine_tasks = [task for task in board_tasks if "루틴" in task.type_names][:12]
        major_tasks = [task for task in board_tasks if "주요" in task.type_names][:12]
        return render_template(
            "dashboard.html",
            routine_tasks=routine_tasks,
            major_tasks=major_tasks,
            task_type_summaries=build_task_type_summaries(tasks, visible_departments),
        )

    @app.get("/tasks")
    @login_required
    def task_list():
        query = visible_task_query(current_user)
        keyword = request.args.get("q", "").strip()
        department_id = request.args.get("department_id", type=int)
        assignee_id = request.args.get("assignee_id", type=int)
        status = request.args.get("status", "")
        task_type = request.args.get("task_type", "")
        cadence = request.args.get("cadence", "")
        start = request.args.get("start", "")
        end = request.args.get("end", "")
        if keyword:
            query = query.where(
                or_(
                    Task.title.ilike(f"%{keyword}%"),
                    Task.content.ilike(f"%{keyword}%"),
                    Task.source_category.ilike(f"%{keyword}%"),
                    Task.source_detail.ilike(f"%{keyword}%"),
                    Task.source_assignees.ilike(f"%{keyword}%"),
                    Task.source_frequency.ilike(f"%{keyword}%"),
                )
            )
        if department_id:
            query = query.where(Task.department_id == department_id)
        if assignee_id:
            query = query.where(Task.assignee_id == assignee_id)
        if status in TASK_STATUSES:
            query = query.where(Task.status == status)
        if task_type in TASK_TYPES:
            query = query.join(TaskClassification).where(TaskClassification.name == task_type)
        if start:
            query = query.where(Task.target_date >= date.fromisoformat(start))
        if end:
            query = query.where(Task.target_date <= date.fromisoformat(end))
        cadence_base_query = query
        cadence_tasks = db.session.scalars(cadence_base_query).unique().all()
        if cadence in WORK_CADENCE_BY_KEY:
            query = query.where(task_cadence_condition(cadence))
        page = max(request.args.get("page", 1, type=int), 1)
        pagination = db.paginate(query.order_by(Task.target_date, Task.id), page=page, per_page=20, error_out=False)
        departments_query = select(Department).where(Department.active.is_(True))
        employees_query = select(Employee).where(
            Employee.status == "재직", Employee.approval_status == "승인완료"
        )
        if current_user.role.data_scope != "all":
            departments_query = departments_query.where(Department.id == current_user.department_id)
            employees_query = employees_query.where(Employee.department_id == current_user.department_id)
        departments = db.session.scalars(departments_query.order_by(Department.name)).all()
        departments.sort(
            key=lambda department: (
                department.id != current_user.department_id,
                department.name,
            )
        )
        all_visible_tasks = db.session.scalars(
            visible_task_query(current_user).order_by(Task.target_date, Task.id)
        ).all()
        journal_selectable_ids = {
            task.id for task in pagination.items if can_write_department(task.department_id)
        }
        major_selectable_ids = {task.id for task in pagination.items if can_edit_task(task)}
        calendar_selectable_ids = {task.id for task in pagination.items if can_edit_task(task)}
        calendar_removable_ids = {
            task.id for task in pagination.items if can_remove_task_calendar(task)
        }
        deletable_task_ids = {task.id for task in pagination.items if can_delete_task(task)}
        selectable_task_ids = (
            journal_selectable_ids | major_selectable_ids | calendar_selectable_ids | deletable_task_ids
        )
        return render_template(
            "tasks.html",
            mode="list",
            pagination=pagination,
            departments=departments,
            employees=db.session.scalars(employees_query.order_by(Employee.name)).all(),
            department_summaries=build_department_summaries(departments, all_visible_tasks),
            cadence_summaries=build_cadence_summaries(cadence_tasks),
            cadence_total=len(cadence_tasks),
            selected_cadence=cadence if cadence in WORK_CADENCE_BY_KEY else "",
            journal_selectable_ids=journal_selectable_ids,
            major_selectable_ids=major_selectable_ids,
            calendar_selectable_ids=calendar_selectable_ids,
            calendar_removable_ids=calendar_removable_ids,
            deletable_task_ids=deletable_task_ids,
            selectable_task_ids=selectable_task_ids,
            can_view_all_tasks=current_user.role.data_scope == "all",
        )

    @app.route("/tasks/new", methods=["GET", "POST"])
    @login_required
    def task_new():
        return task_form_handler(None)

    @app.route("/tasks/<int:task_id>/edit", methods=["GET", "POST"])
    @login_required
    def task_edit(task_id):
        task = get_active_task_or_404(task_id)
        if not can_edit_task(task):
            abort(403)
        return task_form_handler(task)

    @app.post("/tasks/<int:task_id>/delete")
    @login_required
    def task_delete(task_id):
        task = get_active_task_or_404(task_id)
        if not can_delete_task(task):
            abort(403)
        task.deleted_at = utcnow()
        task.deleted_by_id = current_user.id
        audit(
            "TASK_DELETE",
            f"task:{task.id}",
            {
                "title": task.title,
                "department_id": task.department_id,
                "history_preserved": True,
            },
        )
        db.session.commit()
        flash("업무가 삭제되었습니다. 기존 업무일지·회의·감사 이력은 보존됩니다.", "success")
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_list"))

    @app.post("/tasks/bulk-major")
    @login_required
    def task_bulk_major():
        task_ids = {int(item) for item in request.form.getlist("task_ids") if item.isdigit()}
        if not task_ids:
            flash("주요업무로 등록할 업무를 선택해 주세요.", "danger")
            return redirect(url_for("task_list"))
        selected_tasks = db.session.scalars(
            visible_task_query(current_user).where(Task.id.in_(task_ids)).order_by(Task.id)
        ).all()
        if len(selected_tasks) != len(task_ids) or any(not can_edit_task(task) for task in selected_tasks):
            abort(403)
        added_task_ids = []
        for task in selected_tasks:
            if "주요" in task.type_names:
                continue
            was_calendar_included = task.calendar_included
            task.classifications.append(TaskClassification(name="주요"))
            if not was_calendar_included and task.calendar_included:
                task.calendar_registered_by_id = current_user.id
            task.updated_at = utcnow()
            added_task_ids.append(task.id)
        audit(
            "TASK_BULK_MAJOR_ADD",
            "tasks:bulk",
            {
                "requested": len(task_ids),
                "added": len(added_task_ids),
                "task_ids": sorted(task_ids),
                "added_task_ids": added_task_ids,
            },
        )
        db.session.commit()
        flash(
            f"선택한 업무 {len(task_ids)}건 중 {len(added_task_ids)}건을 주요업무에 추가했습니다. "
            "기존 업무 분류는 유지됩니다.",
            "success",
        )
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_list"))

    @app.post("/tasks/bulk-calendar")
    @login_required
    def task_bulk_calendar():
        task_ids = {int(item) for item in request.form.getlist("task_ids") if item.isdigit()}
        if not task_ids:
            flash("일정(캘린더)에 등록할 업무를 선택해 주세요.", "danger")
            return redirect(url_for("task_list"))
        selected_tasks = db.session.scalars(
            visible_task_query(current_user).where(Task.id.in_(task_ids)).order_by(Task.id)
        ).all()
        if len(selected_tasks) != len(task_ids) or any(not can_edit_task(task) for task in selected_tasks):
            abort(403)
        registered_task_ids = []
        for task in selected_tasks:
            if task.calendar_included:
                continue
            task.calendar_excluded = False
            if not task.calendar_auto_included:
                task.calendar_selected = True
            task.calendar_registered_by_id = current_user.id
            task.updated_at = utcnow()
            registered_task_ids.append(task.id)
        audit(
            "TASK_BULK_CALENDAR_ADD",
            "tasks:bulk",
            {
                "requested": len(task_ids),
                "registered": len(registered_task_ids),
                "task_ids": sorted(task_ids),
                "registered_task_ids": registered_task_ids,
            },
        )
        db.session.commit()
        flash(
            f"선택한 업무 {len(task_ids)}건 중 {len(registered_task_ids)}건을 일정(캘린더)에 등록했습니다.",
            "success",
        )
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_list"))

    @app.post("/tasks/bulk-calendar-remove")
    @login_required
    def task_bulk_calendar_remove():
        task_ids = {int(item) for item in request.form.getlist("task_ids") if item.isdigit()}
        if not task_ids:
            flash("일정(캘린더)에서 삭제할 업무를 선택해 주세요.", "danger")
            return redirect(url_for("task_list"))
        selected_tasks = db.session.scalars(
            visible_task_query(current_user).where(Task.id.in_(task_ids)).order_by(Task.id)
        ).all()
        if len(selected_tasks) != len(task_ids) or any(
            not can_remove_task_calendar(task) for task in selected_tasks
        ):
            abort(403)
        removed_task_ids = []
        registered_by_ids = {}
        for task in selected_tasks:
            if not task.calendar_included:
                continue
            registered_by_ids[str(task.id)] = task.calendar_registered_by_id
            task.calendar_selected = False
            task.calendar_excluded = True
            task.calendar_registered_by_id = None
            task.updated_at = utcnow()
            removed_task_ids.append(task.id)
        audit(
            "TASK_BULK_CALENDAR_REMOVE",
            "tasks:bulk",
            {
                "requested": len(task_ids),
                "removed": len(removed_task_ids),
                "task_ids": sorted(task_ids),
                "removed_task_ids": removed_task_ids,
                "calendar_registered_by_ids": registered_by_ids,
                "tasks_preserved": True,
            },
        )
        db.session.commit()
        flash(
            f"선택한 업무 {len(task_ids)}건 중 {len(removed_task_ids)}건을 일정(캘린더)에서 삭제했습니다. "
            "원본 업무와 기존 이력은 유지됩니다.",
            "success",
        )
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_list"))

    @app.post("/tasks/bulk-delete")
    @login_required
    def task_bulk_delete():
        task_ids = {int(item) for item in request.form.getlist("task_ids") if item.isdigit()}
        if not task_ids:
            flash("삭제할 업무를 선택해 주세요.", "danger")
            return redirect(url_for("task_list"))
        selected_tasks = db.session.scalars(
            visible_task_query(current_user).where(Task.id.in_(task_ids)).order_by(Task.id)
        ).all()
        if len(selected_tasks) != len(task_ids) or any(not can_delete_task(task) for task in selected_tasks):
            abort(403)
        deleted_at = utcnow()
        for task in selected_tasks:
            task.deleted_at = deleted_at
            task.deleted_by_id = current_user.id
            audit(
                "TASK_DELETE",
                f"task:{task.id}",
                {
                    "title": task.title,
                    "department_id": task.department_id,
                    "history_preserved": True,
                    "bulk": True,
                },
            )
        audit(
            "TASK_BULK_DELETE",
            "tasks:bulk",
            {"deleted": len(task_ids), "task_ids": sorted(task_ids), "history_preserved": True},
        )
        db.session.commit()
        flash(
            f"선택한 업무 {len(task_ids)}건을 삭제했습니다. 기존 업무일지·회의·감사 이력은 보존됩니다.",
            "success",
        )
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_list"))

    def task_form_handler(task):
        departments_query = select(Department).where(Department.active.is_(True))
        employees_query = select(Employee).where(
            Employee.status == "재직", Employee.approval_status == "승인완료"
        )
        if not current_user.role.allows("task_manage_all"):
            departments_query = departments_query.where(Department.id == current_user.department_id)
            employees_query = employees_query.where(Employee.department_id == current_user.department_id)
        departments = db.session.scalars(departments_query.order_by(Department.name)).all()
        employees = db.session.scalars(employees_query.order_by(Employee.name)).all()
        if request.method == "POST":
            types = [name for name in request.form.getlist("task_types") if name in TASK_TYPES]
            if not types:
                flash("업무 분류를 하나 이상 선택해 주세요.", "danger")
            else:
                department_id = int(request.form["department_id"])
                assignee_id = int(request.form["assignee_id"])
                department = db.session.get(Department, department_id)
                assignee = db.session.get(Employee, assignee_id)
                if (
                    not department
                    or not department.active
                    or not assignee
                    or assignee.status != "재직"
                    or assignee.approval_status != "승인완료"
                ):
                    abort(400)
                if not can_write_department(department_id):
                    abort(403)
                if assignee.department_id != department_id:
                    flash("담당자는 업무 부서에 소속된 임직원만 지정할 수 있습니다.", "danger")
                    return render_template("tasks.html", mode="form", task=task, departments=departments, employees=employees)
                old_status = task.status if task else None
                was_calendar_included = task.calendar_included if task else False
                target = task or Task(created_by_id=current_user.id)
                target.title = request.form.get("title", "").strip()
                target.content = request.form.get("content", "").strip()
                target.work_process = request.form.get("work_process", "").strip() or None
                target.department_id = department_id
                target.assignee_id = assignee_id
                target.start_date = date.fromisoformat(request.form["start_date"])
                target.target_date = date.fromisoformat(request.form["target_date"])
                target.status = request.form.get("status", "진행중") if request.form.get("status") in TASK_STATUSES else "진행중"
                target.progress = min(max(int(request.form.get("progress", 0)), 0), 100)
                target.repeat_cycle = request.form.get("repeat_cycle", "없음") if request.form.get("repeat_cycle") in REPEAT_CYCLES else "없음"
                target.repeat_detail = request.form.get("repeat_detail", "").strip()
                calendar_requested = request.form.get("calendar_included") == "1"
                auto_calendar_requested = any(name in CALENDAR_AUTO_TASK_TYPES for name in types)
                if task is None and auto_calendar_requested and not calendar_requested:
                    target.calendar_excluded = False
                    target.calendar_selected = False
                else:
                    if calendar_requested:
                        target.calendar_excluded = False
                        target.calendar_selected = not auto_calendar_requested
                    else:
                        target.calendar_excluded = (
                            auto_calendar_requested or bool(task and target.calendar_excluded)
                        )
                        target.calendar_selected = False
                if target.status == "완료":
                    target.completed_date = target.completed_date or date.today()
                    target.progress = 100
                else:
                    target.completed_date = None
                if target.target_date < target.start_date:
                    flash("목표일은 착수일보다 빠를 수 없습니다.", "danger")
                elif not target.title:
                    flash("업무 제목을 입력해 주세요.", "danger")
                else:
                    db.session.add(target)
                    db.session.flush()
                    existing_classifications = {item.name: item for item in target.classifications}
                    for name, classification in existing_classifications.items():
                        if name not in types:
                            target.classifications.remove(classification)
                    for name in types:
                        if name not in existing_classifications:
                            target.classifications.append(TaskClassification(name=name))
                    if target.calendar_included and (task is None or not was_calendar_included):
                        target.calendar_registered_by_id = current_user.id
                    elif not target.calendar_included:
                        target.calendar_registered_by_id = None
                    synchronized_task_ids = synchronize_task_work_process(target)
                    if old_status != target.status:
                        db.session.add(
                            TaskStatusLog(
                                task_id=target.id,
                                previous_status=old_status,
                                new_status=target.status,
                                changed_by_id=current_user.id,
                            )
                        )
                    audit(
                        "TASK_UPDATE" if task else "TASK_CREATE",
                        f"task:{target.id}",
                        {
                            "title": target.title,
                            "work_process_updated": bool(target.work_process),
                            "work_process_synchronized_task_ids": synchronized_task_ids,
                            "calendar_selected": target.calendar_selected,
                            "calendar_excluded": target.calendar_excluded,
                            "calendar_registered_by_id": target.calendar_registered_by_id,
                            "calendar_registration": target.calendar_registration_label,
                        },
                    )
                    db.session.commit()
                    if synchronized_task_ids:
                        flash(
                            f"업무가 저장되었으며 동일 원본 업무 {len(synchronized_task_ids)}건의 "
                            "업무 프로세스도 함께 업데이트되었습니다.",
                            "success",
                        )
                    else:
                        flash("업무가 저장되었습니다.", "success")
                    return redirect(url_for("task_detail", task_id=target.id))
        return render_template("tasks.html", mode="form", task=task, departments=departments, employees=employees)

    @app.route("/tasks/<int:task_id>", methods=["GET", "POST"])
    @login_required
    def task_detail(task_id):
        task = get_active_task_or_404(task_id)
        if not can_view_task_detail(task):
            abort(403)
        if request.method == "POST":
            if not can_edit_task(task):
                abort(403)
            content = request.form.get("daily_log", "").strip()
            work_date = date.fromisoformat(request.form.get("work_date", date.today().isoformat()))
            if content:
                db.session.add(TaskDailyLog(task_id=task.id, work_date=work_date, content=content, author_id=current_user.id))
                audit("DAILY_LOG_CREATE", f"task:{task.id}")
                db.session.commit()
                flash("일일업무 기록이 저장되었습니다.", "success")
                return redirect(url_for("task_detail", task_id=task.id))
        status_logs = db.session.scalars(
            select(TaskStatusLog).where(TaskStatusLog.task_id == task.id).order_by(TaskStatusLog.changed_at.desc())
        ).all()
        return render_template(
            "tasks.html",
            mode="detail",
            task=task,
            status_logs=status_logs,
            can_edit=can_edit_task(task),
            can_delete=can_delete_task(task),
        )

    @app.get("/tasks/excel-template")
    @login_required
    def task_excel_template():
        wb = Workbook()
        ws = wb.active
        ws.title = "업무등록"
        headers = ["제목", "내용", "분류", "부서", "담당자 로그인ID", "착수일", "목표일", "상태", "진행률", "반복주기", "캘린더 등록(선택)"]
        ws.append(headers)
        ws.append(["월간 비용 마감", "마감자료 취합", "루틴", current_user.department.name, current_user.login_id, date.today(), date.today() + timedelta(days=7), "진행중", 10, "월간", "Y"])
        ws.append(["월간 핵심 과제", "주요업무는 선택값과 관계없이 자동 등록", "주요", current_user.department.name, current_user.login_id, date.today(), date.today() + timedelta(days=14), "진행중", 0, "없음", "N"])
        ws.freeze_panes = "A2"
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = max(len(str(cell.value or "")) for cell in column) + 3
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="management_task_upload_template.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.post("/tasks/excel-upload")
    @login_required
    def task_excel_upload():
        file = request.files.get("file")
        if not file or not file.filename.lower().endswith(".xlsx"):
            flash(".xlsx 형식의 파일을 선택해 주세요.", "danger")
            return redirect(url_for("task_list"))
        raw = file.read()
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            expected = ["제목", "내용", "분류", "부서", "담당자 로그인ID", "착수일", "목표일", "상태", "진행률", "반복주기"]
            if headers[: len(expected)] != expected:
                raise ValueError("템플릿 헤더가 일치하지 않습니다.")
            calendar_column = headers.index("캘린더 등록(선택)") if "캘린더 등록(선택)" in headers else None
            created = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                department = db.session.scalar(
                    select(Department).where(
                        Department.name == str(row[3]).strip(),
                        Department.active.is_(True),
                    )
                )
                employee = db.session.scalar(select(Employee).where(Employee.login_id == str(row[4]).strip()))
                if not department or not employee:
                    raise ValueError(f"{created + 2}행의 부서 또는 담당자를 찾을 수 없습니다.")
                if (
                    employee.status != "재직"
                    or employee.approval_status != "승인완료"
                    or employee.department_id != department.id
                ):
                    raise ValueError(f"{created + 2}행의 담당자는 해당 부서의 재직자여야 합니다.")
                if not can_write_department(department.id):
                    raise ValueError(f"{created + 2}행은 소속 부서 업무만 등록할 수 있습니다.")
                start_date = row[5].date() if isinstance(row[5], datetime) else (row[5] if isinstance(row[5], date) else date.fromisoformat(str(row[5])))
                target_date = row[6].date() if isinstance(row[6], datetime) else (row[6] if isinstance(row[6], date) else date.fromisoformat(str(row[6])))
                task = Task(
                    title=str(row[0]).strip(),
                    content=str(row[1] or "").strip(),
                    department=department,
                    assignee=employee,
                    start_date=start_date,
                    target_date=target_date,
                    status=str(row[7] or "진행중") if str(row[7] or "진행중") in TASK_STATUSES else "진행중",
                    progress=min(max(int(row[8] or 0), 0), 100),
                    repeat_cycle=str(row[9] or "없음") if str(row[9] or "없음") in REPEAT_CYCLES else "없음",
                    calendar_selected=(
                        str(row[calendar_column] or "").strip().lower()
                        in {"y", "yes", "1", "true", "예", "등록", "표시", "o", "○"}
                        if calendar_column is not None and calendar_column < len(row)
                        else False
                    ),
                    created_by_id=current_user.id,
                )
                uploaded_types = [item.strip() for item in str(row[2]).split("|")]
                type_aliases = {"대표이사수명": "대표이사님 수명업무", "개인": "일반"}
                types = list(
                    dict.fromkeys(
                        type_aliases.get(item, item)
                        for item in uploaded_types
                        if type_aliases.get(item, item) in TASK_TYPES
                    )
                )
                if not types:
                    raise ValueError(f"{created + 2}행의 업무 분류가 올바르지 않습니다.")
                task.classifications = [TaskClassification(name=name) for name in types]
                if task.calendar_included:
                    task.calendar_registered_by_id = current_user.id
                db.session.add(task)
                created += 1
            db.session.add(
                Attachment(
                    related_type="task_excel_upload",
                    related_id=None,
                    filename=file.filename,
                    content_type=file.content_type,
                    file_data=raw,
                    uploader_id=current_user.id,
                )
            )
            audit("TASK_EXCEL_UPLOAD", file.filename, {"created": created})
            db.session.commit()
            flash(f"엑셀에서 업무 {created}건을 등록했습니다.", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"엑셀 등록 실패: {exc}", "danger")
        return redirect(url_for("task_list"))

    @app.get("/calendar")
    @login_required
    def task_calendar():
        today = date.today()
        year = request.args.get("year", today.year, type=int)
        month = request.args.get("month", today.month, type=int)
        if year < 2000 or year > 2100 or month < 1 or month > 12:
            year, month = today.year, today.month
        first = date(year, month, 1)
        last = date(year, month, calendar.monthrange(year, month)[1])
        tasks = db.session.scalars(
            visible_task_query(current_user)
            .where(
                Task.target_date.between(first, last),
                Task.calendar_excluded.is_(False),
                or_(
                    Task.calendar_selected.is_(True),
                    Task.classifications.any(TaskClassification.name.in_(CALENDAR_AUTO_TASK_TYPES)),
                ),
            )
            .order_by(Task.target_date)
        ).all()
        schedules_query = visible_schedule_query(current_user).where(Schedule.schedule_date.between(first, last))
        schedules = db.session.scalars(schedules_query.order_by(Schedule.schedule_date)).all()
        day_items = {}
        for task in tasks:
            day_items.setdefault(task.target_date, []).append(("task", task))
        for schedule in schedules:
            day_items.setdefault(schedule.schedule_date, []).append(("schedule", schedule))
        weeks = calendar.Calendar(firstweekday=6).monthdatescalendar(year, month)
        previous = first - timedelta(days=1)
        next_month = last + timedelta(days=1)
        calendar_rows = sorted(
            [(task.target_date, "task", task) for task in tasks]
            + [(schedule.schedule_date, "schedule", schedule) for schedule in schedules],
            key=lambda row: (row[0], row[1], row[2].id),
        )
        calendar_removable_task_ids = {
            task.id for task in tasks if can_remove_task_calendar(task)
        }
        calendar_deletable_schedule_ids = {
            schedule.id for schedule in schedules if can_delete_schedule(schedule)
        }
        return render_template(
            "calendar.html",
            year=year,
            month=month,
            weeks=weeks,
            day_items=day_items,
            calendar_rows=calendar_rows,
            calendar_removable_task_ids=calendar_removable_task_ids,
            calendar_deletable_schedule_ids=calendar_deletable_schedule_ids,
            previous=previous,
            next_month=next_month,
            can_view_all_schedules=can_view_all_schedules(current_user),
        )

    @app.post("/calendar/tasks/<int:task_id>/remove")
    @login_required
    def task_calendar_remove(task_id):
        task = get_active_task_or_404(task_id)
        if not can_remove_task_calendar(task):
            abort(403)
        registered_by_id = task.calendar_registered_by_id
        task.calendar_selected = False
        task.calendar_excluded = True
        task.calendar_registered_by_id = None
        task.updated_at = utcnow()
        audit(
            "TASK_CALENDAR_REMOVE",
            f"task:{task.id}",
            {
                "title": task.title,
                "calendar_registered_by_id": registered_by_id,
                "task_preserved": True,
                "source": "calendar_page",
            },
        )
        db.session.commit()
        flash("일정을 캘린더에서 삭제했습니다. 원본 업무와 기존 이력은 유지됩니다.", "success")
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_calendar"))

    @app.post("/calendar/schedules/<int:schedule_id>/delete")
    @login_required
    def schedule_calendar_delete(schedule_id):
        schedule = db.session.scalar(
            select(Schedule).where(Schedule.id == schedule_id, Schedule.deleted_at.is_(None))
        )
        if not schedule:
            abort(404)
        if not can_delete_schedule(schedule):
            abort(403)
        schedule.deleted_at = utcnow()
        schedule.deleted_by_id = current_user.id
        audit(
            "SCHEDULE_DELETE",
            f"schedule:{schedule.id}",
            {
                "title": schedule.title,
                "scope": schedule.scope,
                "department_id": schedule.department_id,
                "history_preserved": True,
                "source": "calendar_page",
            },
        )
        db.session.commit()
        flash("등록 일정을 삭제했습니다. 삭제 이력은 감사로그에 보존됩니다.", "success")
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_calendar"))

    @app.route("/meetings", methods=["GET", "POST"])
    @login_required
    def meetings():
        if request.method == "POST":
            task_ids = [int(item) for item in request.form.getlist("task_ids")]
            meeting_date = date.fromisoformat(request.form.get("meeting_date", date.today().isoformat()))
            selected_tasks = db.session.scalars(visible_task_query(current_user).where(Task.id.in_(task_ids))).all() if task_ids else []
            if not selected_tasks:
                flash("회의에 등록할 주요 업무를 선택해 주세요.", "danger")
            else:
                meeting = DailyMeeting(
                    meeting_date=meeting_date,
                    author_id=current_user.id,
                    department_id=current_user.department_id,
                    tasks=selected_tasks,
                )
                db.session.add(meeting)
                db.session.flush()
                audit("MEETING_CREATE", f"meeting:{meeting.id}", {"task_count": len(selected_tasks)})
                db.session.commit()
                return redirect(url_for("meeting_detail", meeting_id=meeting.id))
        candidate_tasks = db.session.scalars(
            visible_task_query(current_user)
            .join(TaskClassification)
            .where(TaskClassification.name == "주요", Task.status != "완료")
            .order_by(Task.target_date)
        ).all()
        meeting_rows = db.session.scalars(select(DailyMeeting).order_by(DailyMeeting.meeting_date.desc(), DailyMeeting.id.desc()).limit(50)).all()
        if current_user.role.data_scope != "all":
            meeting_rows = [row for row in meeting_rows if row.department_id == current_user.department_id or row.author_id == current_user.id]
        return render_template("meetings.html", mode="list", candidate_tasks=candidate_tasks, meetings=meeting_rows)

    @app.get("/meetings/<int:meeting_id>")
    @login_required
    def meeting_detail(meeting_id):
        meeting = db.get_or_404(DailyMeeting, meeting_id)
        if current_user.role.data_scope != "all" and meeting.department_id != current_user.department_id:
            abort(403)
        return render_template("meetings.html", mode="detail", meeting=meeting)

    @app.get("/journals")
    @login_required
    def journals():
        work_date = date.fromisoformat(request.args.get("work_date", date.today().isoformat()))
        employee_id = request.args.get("employee_id", current_user.id, type=int)
        if current_user.role.data_scope == "own":
            employee_id = current_user.id
        elif current_user.role.data_scope == "department":
            employee = db.session.get(Employee, employee_id)
            if not employee or employee.department_id != current_user.department_id:
                employee_id = current_user.id
        employee = db.get_or_404(Employee, employee_id)
        logs = db.session.scalars(
            select(TaskDailyLog)
            .join(Task)
            .where(TaskDailyLog.work_date == work_date, or_(Task.assignee_id == employee_id, TaskDailyLog.author_id == employee_id))
            .order_by(TaskDailyLog.created_at)
        ).all()
        automatic_tasks = db.session.scalars(
            select(Task)
            .join(TaskClassification)
            .where(
                Task.deleted_at.is_(None),
                Task.assignee_id == employee_id,
                TaskClassification.name.in_(["주요", "일반"]),
                Task.start_date <= work_date,
                Task.target_date >= work_date,
            )
            .distinct()
            .order_by(Task.target_date)
        ).all()
        selected_tasks = db.session.scalars(
            select(Task)
            .join(WorkJournalItem, WorkJournalItem.task_id == Task.id)
            .where(
                Task.deleted_at.is_(None),
                WorkJournalItem.work_date == work_date,
                WorkJournalItem.employee_id == employee_id,
            )
            .order_by(Task.target_date, Task.id)
        ).all()
        tasks_by_id = {task.id: task for task in automatic_tasks}
        tasks_by_id.update({task.id: task for task in selected_tasks})
        tasks = sorted(tasks_by_id.values(), key=lambda task: (task.target_date, task.id))
        employees_query = select(Employee).where(
            Employee.status == "재직", Employee.approval_status == "승인완료"
        )
        if current_user.role.data_scope == "department":
            employees_query = employees_query.where(Employee.department_id == current_user.department_id)
        elif current_user.role.data_scope == "own":
            employees_query = employees_query.where(Employee.id == current_user.id)
        employees = db.session.scalars(employees_query.order_by(Employee.name)).all()
        return render_template("journals.html", work_date=work_date, employee=employee, employees=employees, tasks=tasks, logs=logs)

    @app.post("/journals/add-tasks")
    @login_required
    def journal_add_tasks():
        task_ids = {int(item) for item in request.form.getlist("task_ids") if item.isdigit()}
        if not task_ids:
            flash("업무일지에 담을 업무를 선택해 주세요.", "danger")
            return redirect(url_for("task_list"))
        work_date = date.fromisoformat(request.form.get("work_date", date.today().isoformat()))
        selected_tasks = db.session.scalars(
            visible_task_query(current_user).where(Task.id.in_(task_ids)).order_by(Task.id)
        ).all()
        if len(selected_tasks) != len(task_ids) or any(
            not can_write_department(task.department_id) for task in selected_tasks
        ):
            abort(403)
        existing = {
            (task_id, employee_id)
            for task_id, employee_id in db.session.execute(
                select(WorkJournalItem.task_id, WorkJournalItem.employee_id).where(
                    WorkJournalItem.work_date == work_date,
                    WorkJournalItem.task_id.in_(task_ids),
                )
            ).all()
        }
        created = 0
        for task in selected_tasks:
            key = (task.id, task.assignee_id)
            if key in existing:
                continue
            db.session.add(
                WorkJournalItem(
                    task_id=task.id,
                    work_date=work_date,
                    employee_id=task.assignee_id,
                    added_by_id=current_user.id,
                )
            )
            created += 1
        audit(
            "WORK_JOURNAL_BULK_ADD",
            f"work_date:{work_date.isoformat()}",
            {"requested": len(task_ids), "created": created, "task_ids": sorted(task_ids)},
        )
        db.session.commit()
        flash(
            f"선택한 업무 {len(task_ids)}건 중 {created}건을 담당자별 {work_date} 업무일지에 담았습니다.",
            "success",
        )
        return_to = request.form.get("return_to", "")
        if return_to and urlparse(return_to).netloc == "":
            return redirect(return_to)
        return redirect(url_for("task_list"))

    @app.route("/admin/<section>", methods=["GET", "POST"])
    @admin_required("admin")
    def admin(section):
        allowed = {"employees", "roles", "menus", "boards", "schedules", "departments", "audits"}
        if section not in allowed:
            abort(404)
        if request.method == "POST":
            try:
                success_message = handle_admin_post(section)
                db.session.commit()
                flash(success_message or "관리자 설정이 저장되었습니다.", "success")
            except IntegrityError:
                db.session.rollback()
                flash("로그인 ID, 사번 또는 이메일이 이미 등록되어 있습니다.", "danger")
            except ValueError as exc:
                db.session.rollback()
                flash(f"저장할 수 없습니다: {exc}", "danger")
            return redirect(url_for("admin", section=section))
        data = load_admin_data(section)
        return render_template("admin.html", section=section, **data)

    def handle_admin_post(section):
        action = request.form.get("action", "create")
        if section == "employees":
            if action == "approve":
                employee = db.get_or_404(Employee, int(request.form["employee_id"]))
                if employee.approval_status != "승인대기":
                    raise ValueError("승인 대기 중인 계정만 승인할 수 있습니다.")
                if employee.status != "재직":
                    raise ValueError("재직 상태의 계정만 승인할 수 있습니다.")
                employee.approval_status = "승인완료"
                employee.approved_at = utcnow()
                employee.approved_by_id = current_user.id
                employee.failed_login_count = 0
                employee.locked_until = None
                audit(
                    "EMPLOYEE_REGISTRATION_APPROVE",
                    f"employee:{employee.id}",
                    {
                        "login_id": employee.login_id,
                        "department_id": employee.department_id,
                        "role_id": employee.role_id,
                    },
                )
                return f"{employee.name} 임직원의 계정 등록 신청을 승인했습니다."
            if action == "password_reset":
                employee = db.get_or_404(Employee, int(request.form["employee_id"]))
                if employee.status != "재직" or employee.approval_status != "승인완료":
                    raise ValueError("승인 완료된 재직 임직원 계정만 비밀번호를 초기화할 수 있습니다.")
                password = request.form.get("new_password", "")
                confirm_password = request.form.get("confirm_password", "")
                if len(password) < TEMP_PASSWORD_MIN_LENGTH or not any(character.isalpha() for character in password) or not any(
                    character.isdigit() for character in password
                ):
                    raise ValueError(
                        f"임시 비밀번호는 영문과 숫자를 포함하여 {TEMP_PASSWORD_MIN_LENGTH}자 이상이어야 합니다."
                    )
                if password != confirm_password:
                    raise ValueError("임시 비밀번호 확인이 일치하지 않습니다.")
                was_locked = bool(employee.locked_until or employee.failed_login_count)
                employee.set_password(password)
                employee.failed_login_count = 0
                employee.locked_until = None
                employee.must_change_password = True
                audit(
                    "EMPLOYEE_PASSWORD_RESET",
                    f"employee:{employee.id}",
                    {"account_unlocked": was_locked, "must_change_password": True},
                )
                return f"{employee.name} 임직원의 비밀번호를 초기화하고 계정 잠금을 해제했습니다."
            if action == "update":
                employee = db.get_or_404(Employee, int(request.form["employee_id"]))
                login_id = request.form.get("login_id", "").strip()
                if not re.fullmatch(r"[A-Za-z0-9]{4,30}", login_id):
                    raise ValueError("계정 ID는 영문과 숫자만 사용하여 4~30자로 입력해 주세요.")
                name = request.form.get("name", "").strip()
                if not name:
                    raise ValueError("성명을 입력해 주세요.")
                department_id = request.form.get("department_id", type=int)
                department = db.session.get(Department, department_id) if department_id else None
                if not department or not department.active or department.name not in EMPLOYEE_DEPARTMENTS:
                    raise ValueError("등록 가능한 부서(팀)는 보안전산팀, 인사총무팀, 재무운영팀입니다.")
                position = request.form.get("position", "").strip()
                if position not in EMPLOYEE_POSITIONS:
                    raise ValueError("직급은 부서장, 팀장, 팀원 중에서 선택해 주세요.")
                role_id = request.form.get("role_id", type=int)
                role = db.session.get(Role, role_id) if role_id else None
                if not role:
                    raise ValueError("권한을 선택해 주세요.")
                if employee.id == current_user.id and role.name != "관리자":
                    raise ValueError("현재 로그인한 관리자 계정의 관리자 권한은 해제할 수 없습니다.")
                employee.name = name
                employee.employee_no = request.form.get("employee_no", "").strip() or None
                employee.department_id = department_id
                employee.position = position
                employee.email = request.form.get("email", "").strip() or None
                employee.phone = request.form.get("phone", "").strip() or None
                employee.login_id = login_id
                employee.role_id = role_id
                employee.hire_date = (
                    date.fromisoformat(request.form["hire_date"])
                    if request.form.get("hire_date")
                    else None
                )
                audit(
                    "EMPLOYEE_UPDATE",
                    f"employee:{employee.id}",
                    {
                        "fields": [
                            "name",
                            "employee_no",
                            "department_id",
                            "position",
                            "email",
                            "phone",
                            "login_id",
                            "role_id",
                            "hire_date",
                        ]
                    },
                )
                return
            if action == "delete":
                employee = db.get_or_404(Employee, int(request.form["employee_id"]))
                if employee.id == current_user.id:
                    raise ValueError("현재 로그인한 관리자 본인 계정은 삭제할 수 없습니다.")
                if employee.status == "삭제":
                    raise ValueError("이미 삭제된 임직원 계정입니다.")
                employee.status = "삭제"
                employee.termination_date = date.today()
                employee.failed_login_count = 0
                employee.locked_until = None
                audit("EMPLOYEE_DELETE", f"employee:{employee.id}")
                return
            if action in {"terminate", "reactivate"}:
                employee = db.get_or_404(Employee, int(request.form["employee_id"]))
                if employee.approval_status != "승인완료":
                    raise ValueError("승인 대기 계정은 퇴사·재등록 처리할 수 없습니다.")
                if employee.id == current_user.id and action == "terminate":
                    raise ValueError("현재 로그인한 관리자 계정은 종료할 수 없습니다.")
                employee.status = "퇴사" if action == "terminate" else "재직"
                employee.termination_date = date.today() if action == "terminate" else None
                employee.reregistered_at = utcnow() if action == "reactivate" else employee.reregistered_at
                audit("EMPLOYEE_TERMINATE" if action == "terminate" else "EMPLOYEE_REACTIVATE", f"employee:{employee.id}")
                return
            password = request.form.get("password", "")
            login_id = request.form.get("login_id", "").strip()
            if not re.fullmatch(r"[A-Za-z0-9]{4,30}", login_id):
                raise ValueError("계정 ID는 영문과 숫자만 사용하여 4~30자로 입력해 주세요.")
            if len(password) < PASSWORD_MIN_LENGTH or not any(character.isalpha() for character in password) or not any(
                character.isdigit() for character in password
            ):
                raise ValueError(
                    f"초기 비밀번호는 영문과 숫자를 포함하여 {PASSWORD_MIN_LENGTH}자 이상이어야 합니다."
                )
            department_id = request.form.get("department_id", type=int)
            if not department_id:
                raise ValueError("부서(팀)를 선택해 주세요.")
            department = db.session.get(Department, department_id)
            if not department or not department.active or department.name not in EMPLOYEE_DEPARTMENTS:
                raise ValueError("등록 가능한 부서(팀)는 보안전산팀, 인사총무팀, 재무운영팀입니다.")
            position = request.form.get("position", "").strip()
            if position not in EMPLOYEE_POSITIONS:
                raise ValueError("직급은 부서장, 팀장, 팀원 중에서 선택해 주세요.")
            role_id = request.form.get("role_id", type=int)
            role = db.session.get(Role, role_id) if role_id else None
            if not role:
                raise ValueError("권한을 선택해 주세요.")
            name = request.form.get("name", "").strip()
            if not name:
                raise ValueError("성명을 입력해 주세요.")
            employee = Employee(
                name=name,
                employee_no=request.form.get("employee_no", "").strip() or None,
                department_id=department_id,
                position=position,
                email=request.form.get("email", "").strip() or None,
                phone=request.form.get("phone", "").strip() or None,
                login_id=login_id,
                role_id=role.id,
                status="재직",
                hire_date=date.fromisoformat(request.form["hire_date"]) if request.form.get("hire_date") else None,
                password_hash="",
                must_change_password=True,
                approval_status="승인완료",
                approved_at=utcnow(),
                approved_by_id=current_user.id,
            )
            employee.set_password(password)
            db.session.add(employee)
            db.session.flush()
            audit(
                "EMPLOYEE_CREATE",
                f"employee:{employee.id}",
                {"source": "admin", "approved_directly": True, "hire_date_provided": employee.hire_date is not None},
            )
        elif section == "roles":
            permissions = {name: True for name in request.form.getlist("permissions")}
            role = Role(name=request.form["name"].strip(), data_scope=request.form["data_scope"], permissions=permissions)
            db.session.add(role)
            db.session.flush()
            audit("ROLE_CREATE", f"role:{role.id}")
        elif section == "menus":
            menu = Menu(
                name=request.form["name"].strip(),
                url=request.form.get("url", "#").strip(),
                sort_order=int(request.form.get("sort_order", 0)),
                parent_id=int(request.form["parent_id"]) if request.form.get("parent_id") else None,
            )
            menu.roles = db.session.scalars(select(Role).where(Role.id.in_([int(v) for v in request.form.getlist("role_ids")]))).all()
            db.session.add(menu)
            db.session.flush()
            audit("MENU_CREATE", f"menu:{menu.id}")
        elif section == "boards":
            board = Board(
                name=request.form["name"].strip(),
                board_type=request.form.get("board_type", "일반"),
                options={"allow_comments": request.form.get("allow_comments") == "on"},
            )
            board.roles = db.session.scalars(select(Role).where(Role.id.in_([int(v) for v in request.form.getlist("role_ids")]))).all()
            db.session.add(board)
            db.session.flush()
            audit("BOARD_CREATE", f"board:{board.id}")
        elif section == "schedules":
            scope = request.form.get("scope", "전사")
            if scope not in {"전사", "부서", "개인"}:
                raise ValueError("일정 범위가 올바르지 않습니다.")
            department_id = int(request.form["department_id"]) if request.form.get("department_id") else None
            assignee_id = int(request.form["assignee_id"]) if request.form.get("assignee_id") else None
            assignee = db.session.get(Employee, assignee_id) if assignee_id else None
            if assignee_id and (
                not assignee
                or assignee.status != "재직"
                or assignee.approval_status != "승인완료"
            ):
                raise ValueError("승인 완료된 재직 임직원만 일정 담당자로 지정할 수 있습니다.")
            if scope == "부서" and not department_id:
                raise ValueError("부서 일정은 부서를 선택해야 합니다.")
            if scope == "개인" and not assignee_id:
                raise ValueError("개인 일정은 담당자를 선택해야 합니다.")
            if scope == "개인" and not department_id:
                department_id = assignee.department_id
            if assignee and department_id and assignee.department_id != department_id:
                raise ValueError("일정 담당자는 선택한 부서의 임직원이어야 합니다.")
            schedule = Schedule(
                title=request.form["title"].strip(),
                schedule_date=date.fromisoformat(request.form["schedule_date"]),
                scope=scope,
                department_id=department_id,
                assignee_id=assignee_id,
                memo=request.form.get("memo", "").strip(),
                is_holiday=request.form.get("is_holiday") == "on",
                created_by_id=current_user.id,
            )
            db.session.add(schedule)
            db.session.flush()
            audit("SCHEDULE_CREATE", f"schedule:{schedule.id}")
        elif section == "departments":
            department = Department(
                name=request.form["name"].strip(),
                parent_id=int(request.form["parent_id"]) if request.form.get("parent_id") else None,
            )
            db.session.add(department)
            db.session.flush()
            audit("DEPARTMENT_CREATE", f"department:{department.id}")

    def load_admin_data(section):
        common = {
            "roles": db.session.scalars(select(Role).order_by(Role.id)).all(),
            "departments": db.session.scalars(
                select(Department).where(Department.active.is_(True)).order_by(Department.id)
            ).all(),
            "employees": db.session.scalars(
                select(Employee).order_by(
                    case(
                        (
                            and_(
                                Employee.status == "재직",
                                Employee.approval_status == "승인대기",
                            ),
                            0,
                        ),
                        else_=1,
                    ),
                    Employee.status,
                    Employee.name,
                )
            ).all(),
        }
        if section == "employees":
            departments_by_name = {item.name: item for item in common["departments"]}
            common["employee_departments"] = [
                departments_by_name[name] for name in EMPLOYEE_DEPARTMENTS if name in departments_by_name
            ]
            common["employee_positions"] = EMPLOYEE_POSITIONS
        if section == "menus":
            common["menus"] = db.session.scalars(select(Menu).order_by(Menu.sort_order)).all()
        elif section == "boards":
            common["boards"] = db.session.scalars(select(Board).order_by(Board.name)).all()
        elif section == "schedules":
            common["schedules"] = db.session.scalars(
                select(Schedule)
                .where(Schedule.deleted_at.is_(None))
                .order_by(Schedule.schedule_date.desc())
                .limit(100)
            ).all()
        elif section == "audits":
            common["audits"] = db.session.scalars(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(200)).all()
        return common

    @app.errorhandler(403)
    def forbidden(_):
        return render_template("error.html", code=403, message="이 화면을 볼 권한이 없습니다."), 403

    @app.errorhandler(404)
    def not_found(_):
        return render_template("error.html", code=404, message="요청한 화면을 찾을 수 없습니다."), 404

    return app


app = create_app()
run_startup_tasks(app)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8000")))
